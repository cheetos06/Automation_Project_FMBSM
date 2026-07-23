from __future__ import annotations

import argparse
import imaplib
import smtplib
import ssl
import time
import uuid
from email import policy
from email.message import EmailMessage
from email.parser import BytesParser
from pathlib import Path

from dotenv import dotenv_values
from openpyxl import load_workbook


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Send and verify one live balance-cleaner email"
    )
    parser.add_argument("--env", type=Path, required=True)
    parser.add_argument("--workbook", type=Path)
    parser.add_argument(
        "--resume-job-id",
        help="Skip sending and retrieve an existing result by X-FMBSM-Job-ID",
    )
    parser.add_argument("--timeout", type=int, default=3 * 60 * 60)
    parser.add_argument("--output", type=Path, default=Path("balance-email-result"))
    args = parser.parse_args()

    settings = dotenv_values(args.env)
    address = str(settings["GMAIL_ADDRESS"])
    password = str(settings["GMAIL_APP_PASSWORD"]).replace(" ", "")
    token = ""
    if args.resume_job_id:
        print(f"resuming job_id={args.resume_job_id}", flush=True)
    else:
        if args.workbook is None:
            parser.error("--workbook is required unless --resume-job-id is used")
        source = args.workbook.resolve()
        if not source.is_file():
            raise FileNotFoundError(source)
        token = uuid.uuid4().hex[:10]
        subject = f"[balance-cleaner] e2e={token}"
        message = EmailMessage()
        message["From"] = address
        message["To"] = address
        message["Subject"] = subject
        message.set_content(f"Automated end-to-end balance-cleaner test {token}.")
        message.add_attachment(
            source.read_bytes(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=source.name,
        )

        with smtplib.SMTP_SSL(
            str(settings.get("SMTP_HOST") or "smtp.gmail.com"),
            int(settings.get("SMTP_PORT") or 465),
            context=ssl.create_default_context(),
            timeout=60,
        ) as smtp:
            smtp.login(address, password)
            smtp.send_message(message)
        print(f"sent subject={subject}", flush=True)

    deadline = time.monotonic() + args.timeout
    seen_ids: set[bytes] = set()
    while time.monotonic() < deadline:
        with imaplib.IMAP4_SSL(
            str(settings.get("IMAP_HOST") or "imap.gmail.com"),
            int(settings.get("IMAP_PORT") or 993),
            timeout=45,
        ) as imap:
            imap.login(address, password)
            selected = False
            for mailbox in ('"[Gmail]/All Mail"', "INBOX"):
                status, _ = imap.select(mailbox, readonly=True)
                if status == "OK":
                    selected = True
                    break
            if not selected:
                raise RuntimeError("Could not select Gmail All Mail or Inbox")
            if args.resume_job_id:
                status, data = imap.search(
                    None,
                    "HEADER",
                    "X-FMBSM-Job-ID",
                    f'"{args.resume_job_id}"',
                )
            else:
                status, data = imap.search(None, "SUBJECT", f'"{token}"')
            if status != "OK":
                raise RuntimeError("IMAP subject search failed")
            for message_id in reversed(data[0].split()):
                if message_id in seen_ids:
                    continue
                seen_ids.add(message_id)
                status, fetched = imap.fetch(message_id, "(RFC822)")
                if status != "OK":
                    continue
                raw = next((item[1] for item in fetched if isinstance(item, tuple)), None)
                if not raw:
                    continue
                parsed = BytesParser(policy=policy.default).parsebytes(raw)
                reply_type = str(parsed.get("X-FMBSM-Reply-Type", ""))
                print(
                    f"received type={reply_type or 'request'} subject={parsed.get('Subject')}",
                    flush=True,
                )
                body = parsed.get_body(preferencelist=("plain",))
                body_text = body.get_content() if body is not None else ""
                if reply_type == "error":
                    raise RuntimeError(body_text[:2000])
                if reply_type != "result":
                    continue

                args.output.mkdir(parents=True, exist_ok=True)
                workbooks: list[Path] = []
                for part in parsed.iter_attachments():
                    filename = part.get_filename() or "result.bin"
                    target = args.output / filename
                    target.write_bytes(part.get_payload(decode=True) or b"")
                    if target.suffix.lower() == ".xlsx" and target.stat().st_size:
                        workbooks.append(target)
                if not workbooks:
                    raise RuntimeError("Result email had no non-empty Excel attachment")

                record_count = 0
                for result in workbooks:
                    workbook = load_workbook(result, read_only=True, data_only=True)
                    sheet = workbook.active
                    headers = tuple(sheet.cell(1, column).value for column in range(1, 4))
                    if headers != ("account_number", "account_description", "saldo"):
                        workbook.close()
                        raise RuntimeError(f"Unexpected result headers: {headers!r}")
                    record_count += max(0, sheet.max_row - 1)
                    workbook.close()
                if record_count == 0:
                    raise RuntimeError("Result workbook contained no account records")
                print(
                    f"verified workbooks={len(workbooks)} records={record_count} "
                    + "attachments="
                    + ",".join(str(path) for path in workbooks),
                    flush=True,
                )
                return 0
        time.sleep(15)
    target = args.resume_job_id or token
    raise TimeoutError(f"No result email for {target} arrived within {args.timeout} seconds")


if __name__ == "__main__":
    raise SystemExit(main())
