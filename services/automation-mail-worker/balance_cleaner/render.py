from __future__ import annotations

import math
import re
import shutil
import subprocess
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, ImageDraw, ImageFont


OPENXML_SUFFIXES = {".xlsx", ".xlsm"}
LEGACY_SUFFIXES = {".xls", ".xlsb"}


@dataclass(frozen=True)
class Screenshot:
    path: Path
    row_start: int
    row_end: int
    populated_rows: tuple[int, ...]
    first_column: int
    last_column: int
    populated_columns: tuple[int, ...]
    estimated_records: int


@dataclass(frozen=True)
class WorkbookCapture:
    source_path: Path
    workbook_path: Path
    sheet_name: str
    last_populated_row: int
    first_populated_column: int
    last_populated_column: int
    populated_columns: tuple[int, ...]
    header_context: Screenshot | None
    screenshots: tuple[Screenshot, ...]


def normalize_workbook(
    source: Path,
    output_dir: Path,
    *,
    libreoffice_bin: str = "libreoffice",
    timeout_seconds: int = 300,
) -> Path:
    source = source.resolve()
    suffix = source.suffix.lower()
    if suffix in OPENXML_SUFFIXES:
        return source
    if suffix not in LEGACY_SUFFIXES:
        raise ValueError(
            f"Unsupported Excel format {source.suffix or '<none>'}. "
            "Attach .xlsx, .xlsm, .xls, or .xlsb."
        )

    executable = shutil.which(libreoffice_bin) or (
        libreoffice_bin if Path(libreoffice_bin).is_file() else None
    )
    if not executable:
        raise RuntimeError(
            f"LibreOffice executable {libreoffice_bin!r} is unavailable; "
            f"it is required to read {source.suffix} workbooks"
        )
    output_dir.mkdir(parents=True, exist_ok=True)
    completed = subprocess.run(
        [
            str(executable),
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(output_dir),
            str(source),
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout_seconds,
    )
    candidates = sorted(output_dir.glob("*.xlsx"))
    expected = output_dir / f"{source.stem}.xlsx"
    converted = expected if expected.is_file() else (candidates[0] if len(candidates) == 1 else None)
    if completed.returncode != 0 or converted is None:
        detail = (completed.stderr or completed.stdout).strip()[-1500:]
        raise RuntimeError(
            f"LibreOffice could not convert {source.name} to .xlsx"
            + (f": {detail}" if detail else "")
        )
    return converted.resolve()


def capture_workbook(
    source: Path,
    output_dir: Path,
    *,
    rows_per_image: int = 100,
    image_max_width: int = 12000,
    libreoffice_bin: str = "libreoffice",
    last_row_limit: int | None = None,
) -> WorkbookCapture:
    if rows_per_image <= 0:
        raise ValueError("rows_per_image must be positive")
    if image_max_width < 800:
        raise ValueError("image_max_width must be at least 800 pixels")

    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = normalize_workbook(
        source,
        output_dir / "normalized",
        libreoffice_bin=libreoffice_bin,
    )
    try:
        workbook = load_workbook(workbook_path, data_only=False, read_only=False)
        values_workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    except Exception as exc:
        raise RuntimeError(f"Could not open Excel workbook {source.name}: {exc}") from exc

    try:
        selected, bounds = _select_populated_sheet(workbook)
        values_sheet = values_workbook[selected.title]
        min_row, max_row, min_col, max_col, populated_rows = bounds
        if last_row_limit is not None:
            if last_row_limit <= 0:
                raise ValueError("last_row_limit must be positive")
            max_row = min(max_row, last_row_limit)
            populated_rows = {row for row in populated_rows if row <= max_row}
            if not populated_rows:
                raise RuntimeError(
                    f"Workbook {source.name} has no populated rows through row {last_row_limit}"
                )
            max_row = max(populated_rows)
        frozen_rows, frozen_columns = _freeze_table_view(
            selected,
            values_sheet,
            populated_rows,
            min_col,
            max_col,
        )
        if not frozen_rows or not frozen_columns:
            raise RuntimeError(
                f"Workbook {source.name} has no renderable table after removing "
                "empty rows and columns"
            )
        max_row = frozen_rows[-1]
        target_rows = _remove_styled_structure_rows(
            selected,
            values_sheet,
            frozen_rows,
            frozen_columns,
        )
        if not target_rows:
            raise RuntimeError(
                f"Workbook {source.name} has no renderable account rows after "
                "removing styled structural rows"
            )
        account_column = _estimate_account_column(
            selected,
            values_sheet,
            target_rows,
            frozen_columns,
        )
        widths = _column_widths(
            selected,
            values_sheet,
            frozen_columns,
            image_max_width,
        )
        screenshots_dir = output_dir / "screenshots"
        screenshots_dir.mkdir(parents=True, exist_ok=True)
        screenshots: list[Screenshot] = []
        header_rows = _header_rows(
            selected,
            values_sheet,
            frozen_rows,
            frozen_columns,
        )
        header_context: Screenshot | None = None
        if header_rows:
            header_path = screenshots_dir / "header_context.png"
            _render_rows(
                selected,
                values_sheet,
                header_rows,
                frozen_columns,
                widths,
                header_path,
            )
            header_context = Screenshot(
                path=header_path,
                row_start=header_rows[0],
                row_end=header_rows[-1],
                populated_rows=header_rows,
                first_column=frozen_columns[0],
                last_column=frozen_columns[-1],
                populated_columns=frozen_columns,
                estimated_records=0,
            )

        image_number = 1
        for group_start in range(0, len(target_rows), rows_per_image):
            rows = target_rows[group_start : group_start + rows_per_image]
            image_path = screenshots_dir / f"{image_number:04d}.png"
            _render_rows(
                selected,
                values_sheet,
                rows,
                frozen_columns,
                widths,
                image_path,
            )
            screenshots.append(
                Screenshot(
                    path=image_path,
                    row_start=rows[0],
                    row_end=rows[-1],
                    populated_rows=rows,
                    first_column=frozen_columns[0],
                    last_column=frozen_columns[-1],
                    populated_columns=frozen_columns,
                    estimated_records=_estimate_record_count(
                        selected,
                        values_sheet,
                        rows,
                        frozen_columns,
                        account_column,
                    ),
                )
            )
            image_number += 1

        if not screenshots:
            raise RuntimeError(f"Workbook {source.name} did not produce any populated screenshots")
        return WorkbookCapture(
            source_path=source.resolve(),
            workbook_path=workbook_path,
            sheet_name=selected.title,
            last_populated_row=max_row,
            first_populated_column=frozen_columns[0],
            last_populated_column=frozen_columns[-1],
            populated_columns=frozen_columns,
            header_context=header_context,
            screenshots=tuple(screenshots),
        )
    finally:
        workbook.close()
        values_workbook.close()


def _select_populated_sheet(
    workbook,
) -> tuple[Worksheet, tuple[int, int, int, int, set[int]]]:
    candidates: list[
        tuple[tuple[int, int, int], Worksheet, tuple[int, int, int, int, set[int]]]
    ] = []
    for index, sheet in enumerate(workbook.worksheets):
        cells = [cell for cell in sheet._cells.values() if _has_value(cell.value)]  # noqa: SLF001
        if not cells:
            continue
        rows = {cell.row for cell in cells}
        columns = {cell.column for cell in cells}
        bounds = (min(rows), max(rows), min(columns), max(columns), rows)
        score = (len(cells), len(rows), -index)
        candidates.append((score, sheet, bounds))
    if not candidates:
        raise RuntimeError("The workbook contains no populated cells")
    _, sheet, bounds = max(candidates, key=lambda item: item[0])
    return sheet, bounds


def _has_value(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _freeze_table_view(
    sheet: Worksheet,
    values_sheet: Worksheet,
    populated_rows: set[int],
    min_col: int,
    max_col: int,
) -> tuple[tuple[int, ...], tuple[int, ...]]:
    """Choose table rows/columns once, before rendering any screenshot."""
    ordered_rows = tuple(sorted(populated_rows))
    all_columns = tuple(range(min_col, max_col + 1))

    def populated(row: int, column: int) -> bool:
        return _has_value(sheet.cell(row, column).value) or _has_value(
            values_sheet.cell(row, column).value
        )

    # Single-cell title/footer lines must not make a whole spacer column part of
    # the table. Use rows containing at least two values to measure which columns
    # repeatedly participate in the tabular body.
    table_rows = tuple(
        row
        for row in ordered_rows
        if sum(1 for column in all_columns if populated(row, column)) >= 2
    )
    measurement_rows = table_rows or ordered_rows
    minimum_occurrences = max(1, math.ceil(len(measurement_rows) * 0.02))
    columns = tuple(
        column
        for column in all_columns
        if sum(1 for row in measurement_rows if populated(row, column))
        >= minimum_occurrences
    )
    if not columns:
        columns = tuple(
            column
            for column in all_columns
            if any(populated(row, column) for row in ordered_rows)
        )

    # This is the frozen, compact sheet view. Rows that became empty after title
    # and spacer columns were removed are removed here once, not per screenshot.
    rows = tuple(
        row
        for row in ordered_rows
        if any(populated(row, column) for column in columns)
    )
    return rows, columns


def _remove_styled_structure_rows(
    sheet: Worksheet,
    values_sheet: Worksheet,
    rows: tuple[int, ...],
    columns: tuple[int, ...],
) -> tuple[int, ...]:
    """Remove styled hierarchy rows once, before target screenshots are chunked."""

    def populated(row: int, column: int) -> bool:
        return _has_value(sheet.cell(row, column).value) or _has_value(
            values_sheet.cell(row, column).value
        )

    def is_structure_row(row: int) -> bool:
        first = next(
            (sheet.cell(row, column) for column in columns if populated(row, column)),
            None,
        )
        return bool(first is not None and (first.font.bold or first.font.italic))

    return tuple(row for row in rows if not is_structure_row(row))


def _effective_value(
    sheet: Worksheet,
    values_sheet: Worksheet,
    row: int,
    column: int,
) -> Any:
    cell = sheet.cell(row, column)
    cached = values_sheet.cell(row, column)
    return cached.value if cell.data_type == "f" and cached.value is not None else cell.value


def _identifier_key(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _estimate_account_column(
    sheet: Worksheet,
    values_sheet: Worksheet,
    rows: tuple[int, ...],
    columns: tuple[int, ...],
) -> int:
    """Infer the identifier column only to keep Copilot response sizes bounded."""
    scores: list[tuple[int, int, float]] = []
    for column in columns:
        populated = [
            _effective_value(sheet, values_sheet, row, column)
            for row in rows
            if _has_value(_effective_value(sheet, values_sheet, row, column))
        ]
        identifiers: list[str] = []
        for value in populated:
            if isinstance(value, str) and _looks_account_identifier(value):
                identifiers.append(_identifier_key(value))
            elif (
                isinstance(value, (int, float))
                and not isinstance(value, bool)
                and float(value).is_integer()
            ):
                identifiers.append(_identifier_key(value))
        scores.append(
            (
                column,
                len(set(identifiers)),
                len(identifiers) / max(1, len(populated)),
            )
        )
    return max(scores, key=lambda item: (item[1], item[2]))[0]


def _estimate_record_count(
    sheet: Worksheet,
    values_sheet: Worksheet,
    rows: tuple[int, ...],
    columns: tuple[int, ...],
    account_column: int,
) -> int:
    count = 0
    for row in rows:
        account = _effective_value(
            sheet,
            values_sheet,
            row,
            account_column,
        )
        if not _has_value(account) or not _looks_account_identifier(account):
            continue
        if any(
            _looks_numeric(_effective_value(sheet, values_sheet, row, column))
            for column in columns
            if column != account_column
        ):
            count += 1
    return count


def _header_rows(
    sheet: Worksheet,
    values_sheet: Worksheet,
    frozen_rows: tuple[int, ...],
    frozen_columns: tuple[int, ...],
) -> tuple[int, ...]:
    """Return only the pre-table rows to repeat as schema context."""

    def value(row: int, column: int) -> Any:
        cell = sheet.cell(row, column)
        cached = values_sheet.cell(row, column)
        return cached.value if cell.data_type == "f" and cached.value is not None else cell.value

    for index, row in enumerate(frozen_rows):
        row_values = [
            value(row, column)
            for column in frozen_columns
            if _has_value(value(row, column))
        ]
        numeric_like = sum(1 for item in row_values if _looks_numeric(item))
        account_like = len(row_values) >= 2 and _looks_account_identifier(
            row_values[0]
        )
        if account_like or (len(row_values) >= 3 and numeric_like >= 1):
            return frozen_rows[:index]
    return ()


def _looks_numeric(value: Any) -> bool:
    if isinstance(value, bool) or value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    token = str(value).strip().replace("\u00a0", "").replace(" ", "").replace("'", "")
    token = token.strip("()")
    return bool(re.fullmatch(r"[+-]?\d[\d.,]*-?", token))


def _looks_account_identifier(value: Any) -> bool:
    if isinstance(value, bool) or value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    token = str(value).strip()
    return (
        len(token) <= 32
        and any(character.isdigit() for character in token)
        and bool(re.fullmatch(r"[A-Za-z0-9 ._/-]+", token))
    )


def _column_widths(
    sheet: Worksheet,
    values_sheet: Worksheet,
    columns: tuple[int, ...],
    image_max_width: int,
) -> dict[int, int]:
    widths: dict[int, int] = {}
    for column in columns:
        letter = get_column_letter(column)
        configured = float(sheet.column_dimensions[letter].width or 0)
        longest = 0
        for cell in sheet._cells.values():  # noqa: SLF001
            if cell.column != column or not _has_value(cell.value):
                continue
            cached = values_sheet.cell(cell.row, cell.column)
            displayed = _display_value(cell, cached)
            longest = max(longest, *(len(line) for line in displayed.splitlines() or [""]))
        characters = max(8.0, configured, min(float(longest), 48.0))
        widths[column] = min(410, max(68, int(characters * 8 + 14)))

    available = image_max_width - 24
    total = sum(widths.values())
    if total > available:
        scale = available / total
        widths = {column: max(28, int(width * scale)) for column, width in widths.items()}
        total = sum(widths.values())
        if total > available:
            scale = available / total
            widths = {column: max(12, int(width * scale)) for column, width in widths.items()}
    return widths


def _font(bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    names = (
        "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        if bold
        else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf",
    )
    for name in names:
        try:
            return ImageFont.truetype(name, 16)
        except OSError:
            continue
    return ImageFont.load_default()


def _render_rows(
    sheet: Worksheet,
    values_sheet: Worksheet,
    rows: tuple[int, ...],
    columns: tuple[int, ...],
    widths: dict[int, int],
    output_path: Path,
) -> None:
    regular_font = _font(False)
    bold_font = _font(True)
    row_heights: dict[int, int] = {}
    wrapped: dict[tuple[int, int], str] = {}
    for row in rows:
        configured = sheet.row_dimensions[row].height
        height = max(28, int(float(configured) * 96 / 72) if configured else 28)
        for column in columns:
            cell = sheet.cell(row, column)
            cached = values_sheet.cell(row, column)
            value = _display_value(cell, cached)
            lines = _wrap(value, widths[column], bold_font if cell.font.bold else regular_font)
            wrapped[(row, column)] = lines
            line_count = max(1, lines.count("\n") + 1)
            height = max(height, min(260, line_count * 20 + 10))
        row_heights[row] = height

    canvas_width = sum(widths.values()) + 24
    canvas_height = sum(row_heights.values()) + 24
    if canvas_width * canvas_height > 240_000_000:
        raise RuntimeError(
            "A 100-row screenshot would exceed the safe renderer size; "
            "remove unrelated populated columns or excessively long cell text"
        )
    image = Image.new("RGB", (canvas_width, canvas_height), "white")
    draw = ImageDraw.Draw(image)
    x_positions: dict[int, int] = {}
    x = 12
    for column in columns:
        x_positions[column] = x
        x += widths[column]

    y = 12
    for row in rows:
        height = row_heights[row]
        for column in columns:
            cell = sheet.cell(row, column)
            left = x_positions[column]
            right = left + widths[column]
            fill = _cell_color(cell, "fill") or "#FFFFFF"
            font_color = _cell_color(cell, "font") or "#111111"
            draw.rectangle((left, y, right, y + height), fill=fill, outline="#9AA4B2", width=1)
            text = wrapped[(row, column)]
            font = bold_font if cell.font.bold else regular_font
            bbox = draw.multiline_textbbox((0, 0), text, font=font, spacing=2)
            text_height = bbox[3] - bbox[1]
            alignment = str(cell.alignment.horizontal or "general").lower()
            if alignment in {"right", "center"}:
                text_width = bbox[2] - bbox[0]
                text_x = right - text_width - 5 if alignment == "right" else left + (widths[column] - text_width) / 2
            else:
                text_x = left + 5
            vertical = str(cell.alignment.vertical or "center").lower()
            text_y = y + 5 if vertical == "top" else y + max(4, (height - text_height) / 2)
            draw.multiline_text(
                (text_x, text_y),
                text,
                font=font,
                fill=font_color,
                spacing=2,
                align="center" if alignment == "center" else "left",
            )
        y += height
    image.save(output_path, format="PNG", compress_level=6)


def _wrap(value: str, width_pixels: int, font) -> str:
    if not value:
        return ""
    available = max(1, width_pixels - 10)
    lines: list[str] = []
    for original in value.splitlines() or [""]:
        remaining = original
        if not remaining:
            lines.append("")
            continue
        while _text_width(font, remaining) > available:
            low, high = 1, len(remaining)
            while low < high:
                middle = (low + high + 1) // 2
                if _text_width(font, remaining[:middle]) <= available:
                    low = middle
                else:
                    high = middle - 1
            cutoff = max(1, low)
            whitespace = max(
                remaining.rfind(" ", 0, cutoff + 1),
                remaining.rfind("\t", 0, cutoff + 1),
            )
            if whitespace > 0 and remaining[:whitespace].strip():
                line = remaining[:whitespace].rstrip()
                remaining = remaining[whitespace:].lstrip()
            else:
                line = remaining[:cutoff]
                remaining = remaining[cutoff:]
            lines.append(line)
        lines.append(remaining)
    return "\n".join(lines)


def _text_width(font, value: str) -> float:
    getlength = getattr(font, "getlength", None)
    if callable(getlength):
        return float(getlength(value))
    bbox = font.getbbox(value)
    return float(bbox[2] - bbox[0])


def _display_value(cell: Cell, cached: Cell) -> str:
    value = cached.value if cell.data_type == "f" and cached.value is not None else cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S").rstrip(" 00:00:00")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return _format_number(value, cell.number_format)
    return str(value).replace("\x00", "")


def _format_number(value: int | float, number_format: str) -> str:
    if isinstance(value, float) and not math.isfinite(value):
        return str(value)
    section = str(number_format or "General").split(";", 1)[0]
    if "%" in section:
        decimals = _decimal_places(section)
        return f"{float(value) * 100:.{decimals}f}%"
    integer_mask = re.sub(r'"[^"]*"', "", section)
    if re.fullmatch(r"0+", integer_mask) and float(value).is_integer():
        return f"{int(value):0{len(integer_mask)}d}"
    decimals = _decimal_places(section)
    if decimals > 0:
        return f"{float(value):,.{decimals}f}"
    if float(value).is_integer():
        return str(int(value))
    return format(float(value), ".15g")


def _decimal_places(number_format: str) -> int:
    cleaned = re.sub(r'"[^"]*"|\[[^\]]*\]', "", number_format)
    match = re.search(r"[.,](0+)(?:[^0#]|$)", cleaned)
    return len(match.group(1)) if match else 0


def _cell_color(cell: Cell, kind: str) -> str | None:
    if kind == "fill" and cell.fill.fill_type != "solid":
        return None
    color = cell.fill.fgColor if kind == "fill" else cell.font.color
    if color is None or color.type != "rgb" or not color.rgb:
        return None
    raw = str(color.rgb)
    if len(raw) == 8:
        if raw[:2] == "00":
            return None
        raw = raw[2:]
    return f"#{raw}" if re.fullmatch(r"[0-9A-Fa-f]{6}", raw) else None
