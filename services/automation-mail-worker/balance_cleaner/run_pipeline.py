from __future__ import annotations

import json
import math
import os
import re
import sys
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from dotenv import load_dotenv


ROOT = Path(__file__).resolve().parent
PROJECT_DIR = ROOT.parent
FS_REVIEW_DIR = PROJECT_DIR / "fs_review"
if str(FS_REVIEW_DIR) not in sys.path:
    sys.path.insert(0, str(FS_REVIEW_DIR))

from copilot_extract import load_copilot_account  # noqa: E402
from copilot_pool import run_account_pool  # noqa: E402

from .render import (  # noqa: E402
    Screenshot,
    WorkbookCapture,
    capture_workbook,
    select_capture_columns,
)


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def _safe_stem(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._")
    return cleaned[:80] or "balance"


def chunk_screenshots(
    screenshots: tuple[Screenshot, ...],
    batch_size: int,
    max_estimated_records: int = 0,
) -> list[tuple[Screenshot, ...]]:
    effective = min(30, max(1, int(batch_size)))
    record_limit = max(0, int(max_estimated_records))
    batches: list[tuple[Screenshot, ...]] = []
    current: list[Screenshot] = []
    estimated_records = 0
    for screenshot in screenshots:
        estimate = max(0, int(getattr(screenshot, "estimated_records", 0)))
        if current and (
            len(current) >= effective
            or (
                record_limit > 0
                and estimated_records + estimate > record_limit
            )
        ):
            batches.append(tuple(current))
            current = []
            estimated_records = 0
        current.append(screenshot)
        estimated_records += estimate
    if current:
        batches.append(tuple(current))
    return batches


def _batch_prompt(
    base_prompt: str,
    capture: WorkbookCapture,
    batch: tuple[Screenshot, ...],
    batch_index: int,
    batch_count: int,
    schema: dict[str, Any],
    selected_visible_positions: tuple[int, ...],
) -> str:
    labels = "; ".join(
        f"image {index}=Excel rows {item.row_start}-{item.row_end} "
        f"(mechanical candidate-row estimate: "
        f"{max(0, int(getattr(item, 'estimated_records', 0)))})"
        for index, item in enumerate(batch, start=1)
    )
    candidates = [
        {
            "image": image_index,
            "account_number_candidates": list(item.candidate_account_numbers),
        }
        for image_index, item in enumerate(batch, start=1)
    ]
    estimate_note = (
        "The target labels include a mechanical estimate of rows having an "
        "account-like cell and a numeric cell. Use those estimates only as a "
        "navigation and completeness aid when scanning each image. They can "
        "include structural rows, so never extract a title, total, subtotal, "
        "header, or note merely to reach an estimate. "
    )
    compact_schema = json.dumps(
        schema,
        ensure_ascii=False,
        separators=(",", ":"),
    )
    compact_candidates = json.dumps(
        candidates,
        ensure_ascii=False,
        separators=(",", ":"),
    )
    expected_record_count = sum(
        len(item["account_number_candidates"])
        for item in candidates
    )
    return (
        f"{base_prompt.strip()}\n\n"
        f"Batch context: this is batch {batch_index} of {batch_count} for worksheet "
        f"{capture.sheet_name!r}. The batches are consecutive and never overlap. "
        f"{estimate_note}"
        "The complete first screenshot was analyzed separately. Use this fixed "
        "worksheet schema for every target image; original column positions never "
        f"change: {compact_schema}. Do not reinterpret column roles from account "
        "values. "
        "For legibility, every TARGET image contains only these original 1-based "
        "schema column positions, kept in their original left-to-right order: "
        f"{json.dumps(selected_visible_positions)}. Omitted columns are not used "
        "for account_number, account_description, or saldo under the fixed schema. "
        "For navigation and completeness only, a mechanical read of the dedicated "
        "account-number cells found these candidate identifiers in top-to-bottom "
        f"order: {compact_candidates}. Verify every candidate against its visible "
        "row. Include it only when that row visibly has the relevant numeric "
        "balance, and read saldo only from the screenshot. Do not infer any amount "
        "from this list. Do not omit a verified amount-bearing candidate merely "
        "because a nearby blank-account parent or subtotal repeats its amount. "
        f"The candidate lists contain {expected_record_count} amount-bearing "
        "account rows in total. Preserve their top-to-bottom order and return "
        "exactly one object per candidate. Before responding, silently compare the "
        "ordered output account numbers with every per-image candidate list and "
        "correct any omission, duplication, or reordering. Never attach one "
        "candidate's description or saldo to the preceding or following candidate, "
        "even when the rows are visually adjacent; each object's account number, "
        "description, and saldo must come from the same horizontal row. "
        "Extract only rows visible in the attached images and never repeat an account "
        "from a different batch. "
        f"Target image order: {labels}."
    )


def _schema_prompt(base_prompt: str, first_screenshot: Screenshot) -> str:
    candidates = list(first_screenshot.candidate_account_numbers)
    occupancy = [
        {
            "account_number": account_number,
            "nonempty_numeric_column_positions": list(positions),
        }
        for account_number, positions in first_screenshot.candidate_numeric_positions
    ]
    return (
        f"{base_prompt.strip()}\n\n"
        "A mechanical read of the dedicated account-number cells in this first "
        "screenshot found these candidate identifiers in top-to-bottom order: "
        f"{json.dumps(candidates, ensure_ascii=False)}. This list contains no "
        "amounts and is only navigation evidence. Examine the visible row of EVERY "
        "candidate before finalizing the schema. A candidate must not be classified "
        "as an accountless summary when its dedicated account-number cell is visibly "
        "populated. If different candidates place their current balance in different "
        "numeric columns, capture every placement as a separate "
        "valid_account_balance_pattern. The first candidate in this list MUST be the "
        "first worked_visible_example and its exact source column position(s) must "
        "be covered by a valid pattern.\n\n"
        "To prevent visual horizontal shifts across blank cells, here are the "
        "1-based positions of numeric cells that are nonempty on each candidate's "
        "own row. This supplies positions only, never amounts, and does NOT decide "
        "which numeric column is the saldo: "
        f"{json.dumps(occupancy, ensure_ascii=False)}. Use the visible headers to "
        "distinguish current/closing balance positions from prior, comparison, "
        "movement, and variance positions. Every distinct candidate-row placement "
        "must be reconciled with the final schema."
    )


def _extract_schema(
    value: Any,
    visible_column_count: int,
    candidate_numeric_positions: tuple[tuple[str, tuple[int, ...]], ...] = (),
) -> tuple[dict[str, Any], tuple[int, ...], int]:
    if not isinstance(value, dict):
        raise RuntimeError("Copilot header analysis is not a JSON object")
    account_fields = value.get("account_fields")
    saldo = value.get("saldo")
    columns = value.get("columns")
    if not isinstance(account_fields, dict) or not isinstance(saldo, dict):
        raise RuntimeError("Copilot header analysis is missing account_fields or saldo")
    if not isinstance(columns, list):
        raise RuntimeError("Copilot header analysis is missing the complete columns list")

    declared_positions = []
    for item in columns:
        if not isinstance(item, dict):
            raise RuntimeError("Copilot header analysis contains an invalid column item")
        declared_positions.append(_column_position(item.get("position"), "columns.position"))
    expected_positions = list(range(1, visible_column_count + 1))
    if sorted(declared_positions) != expected_positions:
        raise RuntimeError(
            "Copilot header analysis did not enumerate every visible column exactly once"
        )

    account_position = _column_position(
        account_fields.get("account_number_column_position"),
        "account_number_column_position",
    )
    description_position = _column_position(
        account_fields.get("account_description_column_position"),
        "account_description_column_position",
    )
    source_values = saldo.get("source_column_positions")
    if not isinstance(source_values, list) or not source_values:
        raise RuntimeError("Copilot header analysis has no saldo source columns")
    source_positions = [
        _column_position(value, "saldo.source_column_positions")
        for value in source_values
    ]
    required_positions = [account_position, description_position, *source_positions]
    if any(position > visible_column_count for position in required_positions):
        raise RuntimeError(
            "Copilot header analysis references a column outside the first screenshot"
        )
    patterns = saldo.get("valid_account_balance_patterns")
    if not isinstance(patterns, list) or not patterns:
        raise RuntimeError(
            "Copilot header analysis has no valid account balance patterns"
        )
    covered_source_positions: set[int] = set()
    evidence_by_account: dict[str, list[set[int]]] = {}
    for account_number, positions in candidate_numeric_positions:
        evidence_by_account.setdefault(account_number, []).append(set(positions))
    for pattern in patterns:
        if not isinstance(pattern, dict):
            raise RuntimeError("Copilot header analysis contains an invalid balance pattern")
        pattern_values = pattern.get("source_column_positions")
        if not isinstance(pattern_values, list) or not pattern_values:
            raise RuntimeError("Copilot header analysis has a pattern without source columns")
        pattern_positions = {
            _column_position(item, "valid_account_balance_patterns.source_column_positions")
            for item in pattern_values
        }
        covered_source_positions.update(pattern_positions)
        examples = pattern.get("visible_account_examples")
        if not isinstance(examples, list) or not examples:
            raise RuntimeError(
                "Copilot header analysis has a balance pattern without a visible account example"
            )
        for example in examples:
            matches = evidence_by_account.get(str(example), [])
            if matches and not any(pattern_positions & positions for positions in matches):
                raise RuntimeError(
                    "Copilot header analysis horizontally misaligned a balance-pattern example"
                )
    if not set(source_positions).issubset(covered_source_positions):
        raise RuntimeError(
            "Copilot header analysis did not explain every saldo source column "
            "with a valid account pattern"
        )
    if candidate_numeric_positions:
        worked_examples = saldo.get("worked_visible_examples")
        if not isinstance(worked_examples, list) or not worked_examples:
            raise RuntimeError("Copilot header analysis has no worked visible examples")
        first_candidate, first_numeric_positions = candidate_numeric_positions[0]
        first_worked = worked_examples[0]
        if (
            not isinstance(first_worked, dict)
            or str(first_worked.get("account_number")) != first_candidate
        ):
            raise RuntimeError(
                "Copilot header analysis did not use the first candidate as its "
                "first worked example"
            )
        if not set(source_positions) & set(first_numeric_positions):
            raise RuntimeError(
                "Copilot header analysis horizontally misaligned the first account row"
            )
    return value, tuple(sorted(set(required_positions))), account_position


def _column_position(value: Any, label: str) -> int:
    if isinstance(value, bool):
        raise RuntimeError(f"Copilot header analysis has invalid {label}")
    try:
        position = int(value)
    except (TypeError, ValueError) as exc:
        raise RuntimeError(f"Copilot header analysis has invalid {label}") from exc
    if position <= 0:
        raise RuntimeError(f"Copilot header analysis has invalid {label}")
    return position


def _extract_records(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, dict):
        for key in ("accounts", "records", "data", "result"):
            if isinstance(value.get(key), list):
                value = value[key]
                break
    if not isinstance(value, list):
        raise RuntimeError("Copilot response is not a JSON array of accounts")

    records: list[dict[str, Any]] = []
    for index, item in enumerate(value, start=1):
        if not isinstance(item, dict):
            raise RuntimeError(f"Copilot account item {index} is not a JSON object")
        account_number = _account_number(item.get("account_number"))
        description = str(item.get("account_description") or "").strip()
        saldo = _number(item.get("saldo"))
        if not account_number or not description or saldo is None:
            raise RuntimeError(
                f"Copilot account item {index} is missing account_number, "
                "account_description, or numeric saldo"
            )
        records.append(
            {
                "account_number": account_number,
                "account_description": description,
                "saldo": saldo,
            }
        )
    return records


def _one_character_apart(left: str, right: str) -> bool:
    if left == right or abs(len(left) - len(right)) > 1:
        return False
    if len(left) == len(right):
        return sum(a != b for a, b in zip(left, right, strict=True)) == 1
    shorter, longer = (left, right) if len(left) < len(right) else (right, left)
    short_index = 0
    long_index = 0
    differences = 0
    while short_index < len(shorter) and long_index < len(longer):
        if shorter[short_index] == longer[long_index]:
            short_index += 1
            long_index += 1
            continue
        differences += 1
        long_index += 1
        if differences > 1:
            return False
    return True


def _reconcile_account_numbers(
    records: list[dict[str, Any]],
    candidate_account_numbers: tuple[str, ...],
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    candidates = tuple(dict.fromkeys(candidate_account_numbers))
    candidate_set = set(candidates)
    corrections: list[dict[str, str]] = []
    reconciled: list[dict[str, Any]] = []
    for record in records:
        original = str(record["account_number"])
        corrected = original
        if original not in candidate_set and len(original) >= 4:
            matches = [
                candidate
                for candidate in candidates
                if len(candidate) >= 4 and _one_character_apart(original, candidate)
            ]
            if len(matches) == 1:
                corrected = matches[0]
                corrections.append({"from": original, "to": corrected})
        reconciled.append({**record, "account_number": corrected})
    return reconciled, corrections


def _account_number(value: Any) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    token = str(value).strip().replace("\u00a0", "").replace(" ", "").replace("'", "")
    negative_parentheses = token.startswith("(") and token.endswith(")")
    token = token.strip("()")
    if "," in token and "." in token:
        token = (
            token.replace(".", "").replace(",", ".")
            if token.rfind(",") > token.rfind(".")
            else token.replace(",", "")
        )
    elif "," in token:
        token = token.replace(",", ".")
    try:
        number = float(token)
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    return -abs(number) if negative_parentheses else number


def _write_workbook(path: Path, records: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cleaned balance"
    sheet.append(["account_number", "account_description", "saldo"])
    for item in records:
        sheet.append(
            [item["account_number"], item["account_description"], item["saldo"]]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in sheet["A"][1:]:
        cell.number_format = "@"
    for cell in sheet["C"][1:]:
        cell.number_format = '#,##0.00;[Red]-#,##0.00'
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 65
    sheet.column_dimensions["C"].width = 24
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:C{max(1, sheet.max_row)}"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _capture_manifest(capture: WorkbookCapture) -> dict[str, Any]:
    def screenshot_payload(item: Screenshot) -> dict[str, Any]:
        return {
            "path": str(item.path),
            "row_start": item.row_start,
            "row_end": item.row_end,
            "populated_rows": list(item.populated_rows),
            "first_column": item.first_column,
            "last_column": item.last_column,
            "populated_columns": list(item.populated_columns),
            "estimated_records": item.estimated_records,
            "candidate_account_numbers": list(item.candidate_account_numbers),
            "candidate_numeric_positions": [
                {
                    "account_number": account_number,
                    "nonempty_numeric_column_positions": list(positions),
                }
                for account_number, positions in item.candidate_numeric_positions
            ],
        }

    return {
        "source_path": str(capture.source_path),
        "workbook_path": str(capture.workbook_path),
        "sheet_name": capture.sheet_name,
        "last_populated_row": capture.last_populated_row,
        "first_populated_column": capture.first_populated_column,
        "last_populated_column": capture.last_populated_column,
        "populated_columns": list(capture.populated_columns),
        "header_context": (
            screenshot_payload(capture.header_context)
            if capture.header_context is not None
            else None
        ),
        "screenshots": [screenshot_payload(item) for item in capture.screenshots],
    }


def _process_one(
    source: Path,
    workbook_index: int,
    manifest: dict[str, Any],
    settings: dict[str, Any],
    base_prompt: str,
    schema_prompt: str,
) -> Path:
    output_dir = Path(manifest["output_dir"])
    job_id = str(manifest["job_id"])
    work_dir = output_dir / "work" / f"{workbook_index:02d}_{_safe_stem(source.stem)}"
    print(f"[balance-cleaner] rendering {source.name}", flush=True)
    full_capture = capture_workbook(
        source,
        work_dir,
        rows_per_image=int(settings.get("rows_per_image", 100)),
        image_max_width=int(settings.get("image_max_width", 12000)),
        libreoffice_bin=str(settings.get("libreoffice_bin", "libreoffice")),
        last_row_limit=(
            int(settings["_last_row_limit"])
            if settings.get("_last_row_limit") is not None
            else None
        ),
    )
    (work_dir / "full_capture_manifest.json").write_text(
        json.dumps(_capture_manifest(full_capture), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    def schema_operation(_job, selected: dict[str, Any], account: str) -> dict[str, Any]:
        account_config = load_copilot_account(selected)
        from copilot_runtime.runtime import CopilotRuntime

        run_dir = work_dir / "copilot_runs" / "schema" / _safe_stem(account)
        runtime = CopilotRuntime(
            profile_dir=account_config.profile_dir,
            session_dir=account_config.session_dir,
            output_dir=run_dir,
            timeout_seconds=int(selected.get("timeout_seconds", 1200)),
            cleanup=bool(selected.get("cleanup_rendered_images", True)),
            save_private_frames=False,
        )
        started = time.perf_counter()
        metadata = runtime.extract(
            input_path=[full_capture.screenshots[0].path],
            page=1,
            prompt=_schema_prompt(schema_prompt, full_capture.screenshots[0]),
            combine_inputs=True,
            return_metadata=True,
        )
        if not isinstance(metadata, dict):
            raise RuntimeError("Copilot did not return header-analysis metadata")
        schema, selected_positions, account_position = _extract_schema(
            metadata.get("parsed"),
            len(full_capture.populated_columns),
            full_capture.screenshots[0].candidate_numeric_positions,
        )
        print(
            f"[balance-cleaner] schema account={account} "
            f"selected_columns={list(selected_positions)}",
            flush=True,
        )
        return {
            "account": account,
            "elapsed_seconds": round(time.perf_counter() - started, 3),
            "schema": schema,
            "selected_visible_positions": list(selected_positions),
            "account_number_position": account_position,
        }

    schema_pool = run_account_pool([{"kind": "schema"}], settings, schema_operation)
    schema_result = schema_pool.results[0]
    schema = schema_result["schema"]
    selected_visible_positions = tuple(
        int(value) for value in schema_result["selected_visible_positions"]
    )
    account_number_position = int(schema_result["account_number_position"])
    (work_dir / "schema_analysis.json").write_text(
        json.dumps(schema_result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    capture = select_capture_columns(
        full_capture,
        selected_visible_positions,
        account_number_position,
        work_dir / "selected_screenshots",
        image_max_width=int(settings.get("image_max_width", 12000)),
    )
    (work_dir / "capture_manifest.json").write_text(
        json.dumps(_capture_manifest(capture), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    batches = chunk_screenshots(
        capture.screenshots,
        int(settings.get("batch_size", 30)),
        int(settings.get("max_estimated_records_per_request", 0)),
    )
    print(
        f"[balance-cleaner] sheet={capture.sheet_name!r} last_row={capture.last_populated_row} "
        f"screenshots={len(capture.screenshots)} batches={len(batches)}",
        flush=True,
    )

    jobs = [
        {"batch": batch_index, "screenshots": batch}
        for batch_index, batch in enumerate(batches, start=1)
    ]

    def operation(job, selected: dict[str, Any], account: str) -> dict[str, Any]:
        batch_index = int(job["batch"])
        batch = job["screenshots"]
        if len(batch) > 30:
            raise RuntimeError("Copilot balance request exceeds the 30-image limit")
        account_config = load_copilot_account(selected)
        from copilot_runtime.runtime import CopilotRuntime

        run_dir = (
            work_dir
            / "copilot_runs"
            / f"batch_{batch_index:03d}"
            / _safe_stem(account)
        )
        runtime = CopilotRuntime(
            profile_dir=account_config.profile_dir,
            session_dir=account_config.session_dir,
            output_dir=run_dir,
            timeout_seconds=int(selected.get("timeout_seconds", 1200)),
            cleanup=bool(selected.get("cleanup_rendered_images", True)),
            save_private_frames=False,
        )
        started = time.perf_counter()
        metadata = runtime.extract(
            input_path=[item.path for item in batch],
            page=1,
            prompt=_batch_prompt(
                base_prompt,
                capture,
                batch,
                batch_index,
                len(batches),
                schema,
                selected_visible_positions,
            ),
            combine_inputs=True,
            return_metadata=True,
        )
        if not isinstance(metadata, dict):
            raise RuntimeError("Copilot did not return response metadata")
        records = _extract_records(metadata.get("parsed"))
        candidate_account_numbers = tuple(
            account_number
            for screenshot in batch
            for account_number in screenshot.candidate_account_numbers
        )
        records, corrections = _reconcile_account_numbers(
            records,
            candidate_account_numbers,
        )
        print(
            f"[balance-cleaner] batch={batch_index}/{len(batches)} "
            f"account={account} records={len(records)}",
            flush=True,
        )
        return {
            "batch": batch_index,
            "account": account,
            "elapsed_seconds": round(time.perf_counter() - started, 3),
            "records": records,
            "account_number_corrections": corrections,
        }

    pool = run_account_pool(jobs, settings, operation)
    results = list(pool.results)
    by_batch = {
        int(result["batch"]): result
        for result in results
    }

    records: list[dict[str, Any]] = []
    for batch_index in range(1, len(batches) + 1):
        records.extend(by_batch[batch_index]["records"])
    diagnostics = {
        "capture": _capture_manifest(capture),
        "full_capture": _capture_manifest(full_capture),
        "schema_result": schema_result,
        "record_count": len(records),
        "batch_results": results,
        "account_stats": pool.account_stats,
        "schema_account_stats": schema_pool.account_stats,
        "copilot_calls": 1 + len(results),
        "elapsed_seconds": round(
            schema_pool.elapsed_seconds + pool.elapsed_seconds,
            3,
        ),
    }
    (work_dir / "balance_extraction.json").write_text(
        json.dumps(diagnostics, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    suffix = "" if len(manifest["workbook_paths"]) == 1 else f"_{workbook_index:02d}_{_safe_stem(source.stem)}"
    workbook_path = output_dir / f"balance_cleaned_{job_id}{suffix}.xlsx"
    _write_workbook(workbook_path, records)
    # Re-open once so corrupt/partial output is caught before email delivery.
    verification = load_workbook(workbook_path, read_only=True, data_only=True)
    verification.close()
    print(
        f"[balance-cleaner] completed source={source.name} records={len(records)} "
        f"output={workbook_path.name}",
        flush=True,
    )
    return workbook_path


def main() -> int:
    if len(sys.argv) != 3:
        print(
            "Usage: python -m balance_cleaner.run_pipeline manifest.json outputs.json",
            file=sys.stderr,
        )
        return 2
    manifest_path, outputs_path = map(Path, sys.argv[1:3])
    env_file = os.getenv("ENV_FILE")
    load_dotenv(
        dotenv_path=Path(env_file) if env_file else PROJECT_DIR / ".env",
        encoding="utf-8-sig",
    )
    manifest = _read_json(manifest_path)
    settings = _read_json(ROOT / "config" / "settings.json")
    settings["_job_id"] = str(manifest["job_id"])
    prompt_path = ROOT / str(settings["prompt_file"])
    base_prompt = prompt_path.read_text(encoding="utf-8-sig")
    schema_prompt_path = ROOT / str(settings["schema_prompt_file"])
    schema_prompt = schema_prompt_path.read_text(encoding="utf-8-sig")
    sources = [Path(value) for value in manifest["workbook_paths"]]
    if not sources:
        raise RuntimeError("No Excel workbooks were supplied")
    outputs = [
        _process_one(source, index, manifest, settings, base_prompt, schema_prompt)
        for index, source in enumerate(sources, start=1)
    ]
    outputs_path.write_text(
        json.dumps([str(path) for path in outputs], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
