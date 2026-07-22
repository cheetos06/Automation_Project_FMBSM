from __future__ import annotations

import json
import re
import sys
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
PROJECT_DIR = ROOT.parent
FS_REVIEW_DIR = PROJECT_DIR / "fs_review"
if str(FS_REVIEW_DIR) not in sys.path:
    sys.path.insert(0, str(FS_REVIEW_DIR))

from copilot_extract import run_copilot_pdf_extraction  # noqa: E402
from copilot_pool import run_account_pool  # noqa: E402


MONEY_FIELDS = (
    "charges_personnel",
    "salaires_traitements",
    "charges_sociales",
    "autres_charges_salariales",
)
PRIORITY_COLUMNS = (
    "mandate_id",
    "file_name",
    "source_file",
    "status",
    "entity_name",
    "period_end_date",
    "effectif",
    "charges_personnel",
    "salaires_traitements",
    "charges_sociales",
    "autres_charges_salariales",
    "effectif_is_really_zero",
    "effectif_zero_assessment",
    "effectif_justification",
    "effectif_raw_text",
    "effectif_evidence_pages",
    "charges_personnel_raw_text",
    "charges_personnel_evidence_pages",
    "salaires_traitements_raw_text",
    "salaires_traitements_evidence_pages",
    "charges_sociales_raw_text",
    "charges_sociales_evidence_pages",
    "autres_charges_salariales_raw_text",
    "autres_charges_salariales_evidence_pages",
    "confidence",
    "evidence_pages",
    "notes",
    "account_name",
    "elapsed_seconds",
    "error_type",
    "error_message",
    "parsed_json",
)


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    return text or None


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, dict):
        for key in ("normalized_k_eur", "value_k_eur", "value", "amount", "raw_text"):
            parsed = _number(value.get(key))
            if parsed is not None:
                return parsed
        return None
    text = str(value).replace("\xa0", " ")
    matches = re.findall(r"[-+]?\(?\d[\d\s.,']*\)?", text)
    if not matches:
        return None
    raw_token = matches[-1]
    token = raw_token.strip().strip("()").replace("'", "").replace(" ", "")
    negative = "(" in raw_token or token.startswith("-")
    if "," in token and "." in token:
        token = (
            token.replace(".", "").replace(",", ".")
            if token.rfind(",") > token.rfind(".")
            else token.replace(",", "")
        )
    elif "," in token:
        token = (
            token.replace(",", ".")
            if len(token.rsplit(",", 1)[-1]) in {1, 2}
            else token.replace(",", "")
        )
    elif token.count(".") > 1:
        token = token.replace(".", "")
    try:
        number = float(token)
    except ValueError:
        return None
    return -abs(number) if negative else number


def _bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    if text in {"true", "yes", "oui", "1"}:
        return True
    if text in {"false", "no", "non", "0"}:
        return False
    return None


def _mandate_id(path: Path) -> str | None:
    for text in (path.parent.name, path.name):
        match = re.search(r"\b(\d{8,10})\b", text)
        if match:
            return match.group(1)
    return None


def _pages(value: Any) -> str | None:
    if isinstance(value, list):
        return ",".join(str(item) for item in value) or None
    return _clean(value)


def _source_label(pdf: Path, label_root: Path | None) -> str:
    if label_root:
        try:
            return pdf.relative_to(label_root).as_posix()
        except ValueError:
            pass
    return pdf.name


def _flatten(
    parsed: dict[str, Any],
    pdf: Path,
    label: str,
    account: str,
    elapsed: float,
) -> dict[str, Any]:
    raw_values = parsed.get("raw_values") if isinstance(parsed.get("raw_values"), dict) else {}
    row: dict[str, Any] = {
        "mandate_id": _mandate_id(pdf),
        "file_name": pdf.name,
        "source_file": label,
        "status": "ok",
        "entity_name": _clean(parsed.get("entity_name")),
        "period_end_date": _clean(parsed.get("period_end_date")),
        "effectif": _number(parsed.get("effectif")),
        "effectif_is_really_zero": _bool(parsed.get("effectif_is_really_zero")),
        "effectif_zero_assessment": _clean(parsed.get("effectif_zero_assessment")),
        "effectif_justification": _clean(parsed.get("effectif_justification")),
        "confidence": _number(parsed.get("confidence")),
        "evidence_pages": _pages(parsed.get("evidence_pages")),
        "notes": (
            "; ".join(str(item) for item in parsed.get("notes", []))
            if isinstance(parsed.get("notes"), list)
            else _clean(parsed.get("notes"))
        ),
        "account_name": account,
        "elapsed_seconds": round(elapsed, 3),
    }
    for field in MONEY_FIELDS:
        row[field] = _number(parsed.get(field))
    for field in ("effectif", *MONEY_FIELDS):
        raw = raw_values.get(field) if isinstance(raw_values.get(field), dict) else {}
        row[f"{field}_raw_text"] = _clean(raw.get("raw_text"))
        row[f"{field}_evidence_pages"] = _pages(raw.get("evidence_pages"))
    row["parsed_json"] = json.dumps(parsed, ensure_ascii=False, sort_keys=True)
    return row


def _columns(rows: list[dict[str, Any]]) -> list[str]:
    present = {key for row in rows for key in row}
    ordered = [column for column in PRIORITY_COLUMNS if column in present]
    return ordered + sorted(present - set(ordered))


def _write_sheet(workbook: Workbook, name: str, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(name)
    columns = _columns(rows)
    if not columns:
        sheet.append(["status"])
        return
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column) for column in columns])


def _write_workbook(
    path: Path,
    rows: list[dict[str, Any]],
    *,
    elapsed_seconds: float,
    account_stats: dict[str, Any],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(workbook, "effectif_payroll", rows)
    _write_sheet(workbook, "errors", [row for row in rows if row.get("status") != "ok"])
    info = workbook.create_sheet("run_info")
    info.append(["pdf_count", "ok_count", "error_count", "elapsed_seconds", "account_stats_json"])
    ok_count = sum(row.get("status") == "ok" for row in rows)
    info.append(
        [
            len(rows),
            ok_count,
            len(rows) - ok_count,
            round(elapsed_seconds, 3),
            json.dumps(account_stats, ensure_ascii=False),
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    _format(path)


def _format(path: Path) -> None:
    workbook = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in cells[:200]) + 2
            sheet.column_dimensions[get_column_letter(cells[0].column)].width = min(max(width, 10), 48)
    workbook.save(path)


def main() -> int:
    if len(sys.argv) != 3:
        print(
            "Usage: python -m effectif_extract.run_pipeline manifest.json outputs.json",
            file=sys.stderr,
        )
        return 2
    manifest_path, outputs_path = map(Path, sys.argv[1:3])
    manifest = _read_json(manifest_path)
    settings = _read_json(ROOT / "config" / "settings.json")
    settings["_job_id"] = str(manifest["job_id"])
    prompt_path = ROOT / settings["prompt_file"]
    rules_path = ROOT / settings["sector_rules_file"]
    prompt = prompt_path.read_text(encoding="utf-8").replace(
        "{{SECTOR_RULES}}", rules_path.read_text(encoding="utf-8")
    )
    combined_prompt = manifest_path.parent / "effectif_prompt.md"
    combined_prompt.write_text(prompt, encoding="utf-8")
    pdfs = [Path(value) for value in manifest["pdf_paths"]]
    label_root = Path(manifest["label_root"]) if manifest.get("label_root") else None
    runs_dir = Path(manifest["output_dir"]) / "runs"

    def operation(pdf: Path, selected: dict[str, Any], account: str) -> dict[str, Any]:
        started = time.perf_counter()
        try:
            safe_stem = re.sub(r"[^A-Za-z0-9_-]+", "_", pdf.stem)[:60]
            parsed = run_copilot_pdf_extraction(
                pdf,
                combined_prompt,
                runs_dir / account / f"{safe_stem}_{time.time_ns()}",
                selected,
                tag="effectif",
            )
            return _flatten(
                parsed,
                pdf,
                _source_label(pdf, label_root),
                account,
                time.perf_counter() - started,
            )
        except Exception as exc:
            print(
                f"[effectif] ignored error account={account} file={pdf.name}: {exc}",
                flush=True,
            )
            return {
                "mandate_id": _mandate_id(pdf),
                "file_name": pdf.name,
                "source_file": _source_label(pdf, label_root),
                "status": "error",
                "error_type": type(exc).__name__,
                "error_message": str(exc)[:2000],
                "account_name": account,
                "elapsed_seconds": round(time.perf_counter() - started, 3),
            }

    pool = run_account_pool(pdfs, settings, operation)
    rows = pool.results
    output_dir = Path(manifest["output_dir"])
    workbook_path = output_dir / f"effectif_payroll_{manifest['job_id']}.xlsx"
    _write_workbook(
        workbook_path,
        rows,
        elapsed_seconds=pool.elapsed_seconds,
        account_stats=pool.account_stats,
    )
    outputs_path.write_text(json.dumps([str(workbook_path)], indent=2), encoding="utf-8")
    ok = sum(row.get("status") == "ok" for row in rows)
    print(
        f"[effectif] finished pdfs={len(rows)} ok={ok} errors={len(rows)-ok} "
        f"output={workbook_path}",
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
