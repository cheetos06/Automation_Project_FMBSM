from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import aspose.pdf as pdf
from cryptography import x509
from cryptography.x509.oid import NameOID
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

DATE_FMT = "%d/%m/%Y"
logger = logging.getLogger(__name__)


def _format_date(value) -> Tuple[str, str, Optional[datetime]]:
    """Normalize to dd/mm/yyyy and hh:mm:ss."""
    if isinstance(value, datetime):
        return value.strftime(DATE_FMT), value.strftime("%H:%M:%S"), value

    if isinstance(value, str):
        for fmt in (
            "%Y-%m-%dT%H:%M:%S%z",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y",
        ):
            try:
                dt = datetime.strptime(value, fmt)
                return dt.strftime(DATE_FMT), dt.strftime("%H:%M:%S"), dt
            except ValueError:
                continue

    return str(value), "", None


def extract_digital_signatures(pdf_path: str) -> List[Dict[str, Optional[str]]]:
    """Return signer name and date for each digital signature embedded in pdf_path."""
    pdf_sign = pdf.facades.PdfFileSignature()
    pdf_sign.bind_pdf(str(pdf_path))

    if not pdf_sign.contains_signature():
        return []

    signatures: List[Dict[str, Optional[str]]] = []
    for sig_name in pdf_sign.get_sign_names(False):
        date_raw = pdf_sign.get_date_time(sig_name)
        date_str, time_str, parsed_dt = _format_date(date_raw)
        signer = pdf_sign.get_signer_name(sig_name)

        cert_stream = pdf_sign.extract_certificate(sig_name)
        if cert_stream:
            cert_content = cert_stream.read()
            try:
                cert = x509.load_der_x509_certificate(cert_content)
                cn = cert.subject.get_attributes_for_oid(NameOID.COMMON_NAME)
                if cn and (not signer or signer == "None"):
                    signer = cn[0].value
            except Exception:
                pass

        if not signer or signer == "None":
            signer = sig_name

        signatures.append(
            {
                "signer": signer,
                "date": date_str,
                "parsed_date": parsed_dt,
                "time": time_str,
            }
        )

    return signatures


def _sort_signatures(signatures: List[Dict[str, Optional[str]]]) -> List[Dict[str, Optional[str]]]:
    """Sort signatures: DocuSign signers first, otherwise by most recent date first."""

    def sort_key(sig: Dict[str, Optional[str]]):
        name = (sig.get("signer") or "").lower()
        docusign_flag = 0 if "docusign" in name else 1
        dt = sig.get("parsed_date")
        ord_val = dt.toordinal() if isinstance(dt, datetime) else -1
        return (docusign_flag, -ord_val, name)

    return sorted(signatures, key=sort_key)


def collect_signatures_for_pdfs(
    pdf_paths: Sequence[Path],
    *,
    label_root: Path | None = None,
) -> List[Dict[str, object]]:
    """Collect signatures for multiple PDFs; always include PDFs even if empty."""
    results: List[Dict[str, object]] = []
    for pdf_path in pdf_paths:
        try:
            signatures = _sort_signatures(extract_digital_signatures(str(pdf_path)))
        except Exception:
            logger.exception("Error reading PDF signatures from %s", pdf_path)
            signatures = []
        results.append({"file": _file_label(pdf_path, label_root), "path": str(pdf_path), "signatures": signatures})
    return results


def write_wide_dates(entries: Sequence[Dict[str, object]], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dates"

    max_sigs = max((len(e["signatures"]) for e in entries), default=0)
    headers = ["File"]
    for i in range(max_sigs):
        headers.extend([f"Date {i + 1}", f"Time {i + 1}"])
    ws.append(headers)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for entry in entries:
        sigs = entry["signatures"]
        row = [entry["file"]]
        for sig in sigs:
            row.extend([sig.get("date", ""), sig.get("time", "")])
        ws.append(row)
        if not sigs:
            for cell in ws[ws.max_row]:
                cell.fill = yellow_fill

    ws.column_dimensions["A"].width = 55
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def write_long_signers(entries: Sequence[Dict[str, object]], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Signers"

    ws.append(["File", "Signer", "Date", "Time"])
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    current_row = 2
    for entry in entries:
        sigs = entry["signatures"]
        if not sigs:
            ws.append([entry["file"], "", "", ""])
            for cell in ws[current_row]:
                cell.fill = yellow_fill
            current_row += 1
            continue

        start_row = current_row
        for sig in sigs:
            ws.append([entry["file"], sig.get("signer", ""), sig.get("date", ""), sig.get("time", "")])
            current_row += 1
        end_row = current_row - 1
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.cell(row=start_row, column=1).alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def create_excel_outputs(
    pdf_paths: Sequence[Path],
    output_dir: Path,
    job_id: str,
    *,
    label_root: Path | None = None,
) -> List[Path]:
    entries = collect_signatures_for_pdfs(pdf_paths, label_root=label_root)
    output_dir.mkdir(parents=True, exist_ok=True)
    dates_path = output_dir / f"digital_sign_dates_{job_id}.xlsx"
    signers_path = output_dir / f"digital_sign_signers_{job_id}.xlsx"
    return [
        write_wide_dates(entries, dates_path),
        write_long_signers(entries, signers_path),
    ]


def _file_label(pdf_path: Path, label_root: Path | None) -> str:
    if label_root:
        try:
            return pdf_path.resolve(strict=False).relative_to(label_root.resolve(strict=False)).as_posix()
        except ValueError:
            pass
    return pdf_path.name
