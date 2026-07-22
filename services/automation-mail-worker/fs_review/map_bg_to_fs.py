"""Map a standardized French general ledger (BG) to extracted FS line labels.

Prediction inputs are only bg_standardized.xlsx and fs_extract_N.json. FAST
files and historical ticket mappings are never opened. Runtime classification
uses explicit French-PCG accounting rules and strict amount reconciliation.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from fs_mapping import build_fs_mapping_rows, write_fs_mapping
from pcg_rules import account_digits
from rule_based_mapper import PCGMapper


OUTPUT_HEADERS = ["bg account", "libelle", "montant", "mapping FS"]
TOKEN_RE = re.compile(r"[a-z0-9]+")

ACCOUNT_HEADERS = {"compte", "account number", "account", "bg account"}
LABEL_HEADERS = {"libelle", "account label", "label"}
FALLBACK_LABEL_HEADERS = {"libelle initial", "account number 2"}
AMOUNT_HEADERS = {"montant", "balance", "amount n", "solde"}
DEBIT_HEADERS = {"debit"}
CREDIT_HEADERS = {"credit"}

STOPWORDS = {
    "a",
    "au",
    "aux",
    "d",
    "de",
    "des",
    "du",
    "en",
    "et",
    "l",
    "la",
    "le",
    "les",
    "par",
    "pour",
    "sur",
}

STEMS = {
    "acomp": "acompte",
    "acpt": "acompte",
    "amort": "amortissement",
    "approv": "approvisionnement",
    "avance": "avance",
    "brevet": "brevet",
    "capital": "capital",
    "charge": "charge",
    "client": "client",
    "concession": "concession",
    "constat": "constate",
    "construct": "construction",
    "cotisation": "cotisation",
    "creance": "creance",
    "depreci": "depreciation",
    "dette": "dette",
    "developp": "developpement",
    "disponib": "disponibilite",
    "emprunt": "emprunt",
    "exception": "exceptionnel",
    "exploit": "exploitation",
    "financ": "financier",
    "fourniss": "fournisseur",
    "immob": "immobilisation",
    "impot": "impot",
    "incorpor": "incorporel",
    "interet": "interet",
    "marchand": "marchandise",
    "mater": "matiere",
    "produc": "production",
    "provis": "provision",
    "reserve": "reserve",
    "resultat": "resultat",
    "salaire": "salaire",
    "social": "social",
    "stock": "stock",
    "subvention": "subvention",
    "tax": "taxe",
    "vend": "vente",
}

SUMMARY_TERMS = (
    "total ",
    "produits exploitation",
    "charge exploitation",
    "resultat exploitation",
    "resultat financier",
    "resultat courant",
    "resultat exceptionnel",
    "benefice ou perte",
)

EXACT_SUMMARY_LABELS = {
    "total",
    "actif circulant",
    "actif immobilisation",
    "capital propre",
    "capitaux propres",
    "dette",
    "provision",
    "chiffre affaires net",
    "chiffres affaires nets",
    "montant net chiffre affaires",
    "situation nette",
    "produits financier",
    "charge financier",
    "benefice exercice avant impot participation",
    "benefice net exercice",
}


@dataclass
class BGRow:
    account: Any
    label: str
    amount: float
    mapping: str = ""


@dataclass
class FSLine:
    label: str
    amount: float | None
    amount_n_1: float | None
    gross: float | None
    depreciation: float | None
    normalized: str
    summary: bool
    scope: str = ""
    statement: str = ""
    statement_family: str = ""
    result_section: str = ""
    page: int | None = None
    canonical_id: str = ""


LEGACY_PRESENTATION_MARKERS = (
    "transferts de charges",
    "operations de gestion",
    "operations en capital",
    "valeurs comptables elements actif cedes",
)

MODERN_PRESENTATION_MARKERS = (
    "produits de cession immobilisations incorporelles corporelles",
    "valeurs comptables immobilisations incorporelles corporelles cedees",
)


def strip_accents(value: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(char)
    )


def normalized_words(value: Any) -> list[str]:
    text = strip_accents(str(value or "")).lower().replace("&", " et ")
    raw_tokens = TOKEN_RE.findall(text)
    result: list[str] = []
    for token in raw_tokens:
        if token in STOPWORDS:
            continue
        stemmed = token
        for prefix, replacement in STEMS.items():
            if token.startswith(prefix):
                stemmed = replacement
                break
        result.append(stemmed)
    return result


def normalize_text(value: Any) -> str:
    return " ".join(normalized_words(value))


def is_summary_label(normalized: str) -> bool:
    return normalized in EXACT_SUMMARY_LABELS or any(
        term in normalized for term in SUMMARY_TERMS
    )


def numeric(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if not text:
        return None
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    elif "," in text and "." in text:
        text = text.replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def header_key(value: Any) -> str:
    return normalize_text(value).replace("_", " ")


def find_header(values: list[Any], accepted: set[str]) -> int | None:
    for index, value in enumerate(values):
        if header_key(value) in accepted:
            return index
    return None


def load_bg(path: Path) -> list[BGRow]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    all_rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not all_rows:
        return []

    header_index = None
    columns: tuple[
        int | None, int | None, int | None, int | None, int | None, int | None
    ]
    for index, row in enumerate(all_rows[:20]):
        values = list(row)
        account_index = find_header(values, ACCOUNT_HEADERS)
        label_index = find_header(values, LABEL_HEADERS)
        amount_index = find_header(values, AMOUNT_HEADERS)
        debit_index = find_header(values, DEBIT_HEADERS)
        credit_index = find_header(values, CREDIT_HEADERS)
        fallback_label_index = find_header(values, FALLBACK_LABEL_HEADERS)
        has_amount = (
            amount_index is not None
            or debit_index is not None
            or credit_index is not None
        )
        if account_index is not None and label_index is not None and has_amount:
            header_index = index
            columns = (
                account_index,
                label_index,
                amount_index,
                debit_index,
                credit_index,
                fallback_label_index,
            )
            break
    if header_index is None:
        raise ValueError(f"Could not identify BG columns in {path}")

    (
        account_index,
        label_index,
        amount_index,
        debit_index,
        credit_index,
        fallback_label_index,
    ) = columns
    rows: list[BGRow] = []
    for raw in all_rows[header_index + 1 :]:
        account = raw[account_index] if account_index < len(raw) else None
        label = raw[label_index] if label_index < len(raw) else None
        if label in (None, "") and fallback_label_index is not None:
            label = (
                raw[fallback_label_index]
                if fallback_label_index < len(raw)
                else None
            )
        if account in (None, ""):
            continue
        if amount_index is not None:
            amount = numeric(raw[amount_index] if amount_index < len(raw) else None)
        else:
            debit = (
                numeric(raw[debit_index] if debit_index is not None else None) or 0.0
            )
            credit = (
                numeric(raw[credit_index] if credit_index is not None else None)
                or 0.0
            )
            amount = debit - credit
        rows.append(BGRow(account, str(label or "").strip(), amount or 0.0))
    return rows


def json_lines(data: Any) -> list[dict[str, Any]]:
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if isinstance(data, dict):
        for key in ("lines", "lignes", "data", "items", "fs_lines"):
            if isinstance(data.get(key), list):
                return [item for item in data[key] if isinstance(item, dict)]
    return []


def statement_family(value: Any) -> str:
    """Normalize source statement names to accounting presentation families."""

    normalized = normalize_text(value)
    if "actif" in normalized and "passif" not in normalized:
        return "asset"
    if "passif" in normalized:
        return "liability"
    if any(
        marker in normalized
        for marker in ("compte resultat", "resultat", "produits charges")
    ):
        return "result"
    return ""


def infer_statement_families(lines: list[FSLine]) -> None:
    """Infer sections for legacy flat extracts using their statutory order.

    Current canonical extracts carry an explicit statement name. Historical
    service-desk fixtures are flat lists, but preserve the PCG order: assets,
    liabilities/equity, then the income statement. The inference is deliberately
    based on structural anchors, never on ticket identifiers or amounts.
    """

    if not lines:
        return

    # ANC 2016-03 places SCPI assets and negative passifs in one combined
    # "état du patrimoine".  Preserve that statutory structure instead of
    # treating the full page as an ordinary asset statement.
    for line in lines:
        if "etat patrimoine" not in normalize_text(line.statement):
            continue
        label = line.normalized
        line.statement_family = (
            "liability"
            if label.startswith(("dette", "dettes", "total iv", "capitaux propre"))
            else "asset"
        )

    if all(line.statement_family for line in lines):
        return

    def is_liability_start(label: str) -> bool:
        if "non appele" in label or "non verse" in label:
            return False
        return (
            label == "capital"
            or label.startswith("capital social")
            or label.startswith("capital dont verse")
            or label.startswith("fonds associatif")
            or label.startswith("fonds propre")
        )

    def is_result_start(label: str) -> bool:
        return label.startswith("chiffre affaires") or any(
            marker in label
            for marker in (
                "vente marchandise",
                "ventes marchandise",
                "production vente",
                "chiffre affaires net",
                "chiffres affaires nets",
                "montant net chiffre affaires",
                "produits exploitation",
                "charges exploitation",
                "consommations exercice",
                "achats marchandises",
                "autres achats charges externes",
            )
        )

    liability_start = next(
        (
            index
            for index, line in enumerate(lines)
            if not line.statement_family and is_liability_start(line.normalized)
        ),
        None,
    )
    first_result_start = next(
        (
            index
            for index, line in enumerate(lines)
            if not line.statement_family and is_result_start(line.normalized)
        ),
        None,
    )
    result_search_start = (liability_start + 1) if liability_start is not None else 0
    result_start = next(
        (
            index
            for index, line in enumerate(lines[result_search_start:], result_search_start)
            if not line.statement_family and is_result_start(line.normalized)
        ),
        None,
    )

    # A few historical extracts concatenate the income statement before the
    # balance sheet. Preserve that explicit block rather than assuming that
    # every flat list is asset/passif/result ordered.
    if (
        first_result_start is not None
        and liability_start is not None
        and first_result_start < liability_start
    ):
        balance_total_start = next(
            (
                index
                for index, line in enumerate(lines[first_result_start:], first_result_start)
                if "total actif" in line.normalized
                or "total passif" in line.normalized
            ),
            liability_start,
        )
        for index, line in enumerate(lines):
            if line.statement_family:
                continue
            if index < balance_total_start:
                line.statement_family = "result"
            elif "actif" in line.normalized and "passif" not in line.normalized:
                line.statement_family = "asset"
            else:
                line.statement_family = "liability"
        return

    # A result-only extract has no equity anchor. Its first recognizable result
    # line establishes the family for the entire supplied statement.
    if liability_start is None and result_start is not None:
        liability_start = 0

    for index, line in enumerate(lines):
        if line.statement_family:
            continue
        if result_start is not None and index >= result_start:
            line.statement_family = "result"
        elif liability_start is not None and index >= liability_start:
            line.statement_family = "liability"
        else:
            line.statement_family = "asset"


def infer_result_sections(lines: list[FSLine]) -> None:
    """Infer operating/financial/exceptional/tax blocks from PCG headings."""

    section = ""
    for line in lines:
        if line.statement_family != "result":
            continue
        if line.result_section:
            section = line.result_section
            continue
        label = line.normalized
        leading_account = label.split(" ", 1)[0]
        if leading_account.startswith(("73", "74", "75")):
            section = "operating"
        elif leading_account.startswith("76"):
            section = "financial"
        elif leading_account.startswith("77"):
            section = "exceptional"
        elif any(
            marker in label
            for marker in (
                "exceptionnel",
                "exceptionnelle",
                "prod except",
                "prod excep",
                "charge except",
                "charges except",
                "charg except",
            )
        ):
            section = "exceptional"
        elif any(
            marker in label
            for marker in (
                "produit financier",
                "produits financier",
                "charge financier",
                "charges financier",
                "resultat financier",
                "participation financier",
                "autres valeurs mobilieres creance actif immobilisation",
                "autres interet produits assimiles",
                "interet produit assimile",
                "interet charge assimile",
                "difference positive change",
                "difference negative change",
                "differences positives change",
                "differences negatives change",
            )
        ) or (
            label.startswith("participation") and "salarie" not in label
        ):
            section = "financial"
        elif any(
            marker in label
            for marker in (
                "impot benefice",
                "participation salarie",
                "participation salaries",
            )
        ):
            section = "tax"
        elif any(
            marker in label
            for marker in (
                "exploitation",
                "activite",
                "vente marchandise",
                "ventes marchandise",
                "production vente",
                "achat marchandise",
                "achats marchandise",
                "salaire",
                "cotisation social",
            )
        ):
            section = "operating"
        line.result_section = section


def load_fs(path: Path) -> list[FSLine]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    source_lines = json_lines(data)
    scpi_extract = any(
        "etat patrimoine"
        in normalize_text(item.get("statement") or item.get("etat") or "")
        for item in source_lines
    )
    scpi_parent_labels = {
        "immobilisation locatives",
        "provision liees placements immobilisation",
        "creance",
        "valeurs placement disponibilite",
        "produits immobilisation i",
        "charge activite immobilisation ii",
        "produits exploitation i",
        "charge exploitation ii",
        "produits financier i",
        "charge financier ii",
        "produits exceptionnel i",
        "charge exceptionnel ii",
    }
    result: list[FSLine] = []
    for item in source_lines:
        label = item.get("libelle") or item.get("label") or item.get("name")
        if not label:
            continue
        normalized = normalize_text(label)
        source_statement = str(item.get("statement") or item.get("etat") or "")
        source_section = str(item.get("section") or item.get("result_section") or "")
        explicit_total = bool(item.get("is_total", item.get("isTotal", False)))
        forced_detail = scpi_extract and normalized in {"charge exploitation societe"}
        scpi_parent = scpi_extract and normalized in scpi_parent_labels
        result.append(
            FSLine(
                label=str(label).strip(),
                amount=numeric(item.get("montant_n", item.get("amount_n"))),
                amount_n_1=numeric(
                    item.get("montant_n_1", item.get("amount_n_1"))
                ),
                gross=numeric(item.get("brut", item.get("gross"))),
                depreciation=numeric(
                    item.get("amortissement", item.get("depreciation"))
                ),
                normalized=normalized,
                summary=(explicit_total or is_summary_label(normalized) or scpi_parent)
                and not forced_detail,
                scope=str(item.get("scope") or ""),
                statement=source_statement,
                statement_family=statement_family(source_statement),
                result_section=normalize_text(source_section),
                page=(
                    int(item["page"])
                    if str(item.get("page", "")).strip().isdigit()
                    else None
                ),
                canonical_id=str(item.get("canonical_id") or ""),
            )
        )
    infer_statement_families(result)
    infer_result_sections(result)
    return result


def detect_presentation_regime(
    rows: list[BGRow],
    fs_lines: list[FSLine],
    reporting_year: int | None,
) -> tuple[str, list[str]]:
    """Detect legacy/current presentation from source evidence, then year."""

    warnings: list[str] = []
    labels = " | ".join(line.normalized for line in fs_lines)
    legacy_score = sum(marker in labels for marker in LEGACY_PRESENTATION_MARKERS)
    modern_score = sum(marker in labels for marker in MODERN_PRESENTATION_MARKERS)
    nonzero_codes = {
        account_digits(row.account)
        for row in rows
        if abs(row.amount) > 0.01
    }
    if any(code.startswith(("675", "775", "791")) for code in nonzero_codes):
        legacy_score += 1
    if any(code.startswith(("657", "757")) for code in nonzero_codes):
        modern_score += 1

    if legacy_score > modern_score:
        regime = "legacy"
    elif modern_score > legacy_score:
        regime = "modern"
    elif reporting_year is not None:
        regime = "modern" if reporting_year >= 2025 else "legacy"
    else:
        regime = "modern"
        warnings.append(
            "Presentation regime was not identifiable; defaulted to current PCG. "
            "Pass --year when processing an older dossier."
        )

    if reporting_year is not None and reporting_year >= 2025 and regime == "legacy":
        warnings.append(
            f"{reporting_year} dossier uses legacy presentation signatures; "
            "verify whether this is a tax-form layout, an old template, or a "
            "period/extraction error"
        )
    return regime, warnings


def fs_quality_warnings(fs_lines: list[FSLine]) -> list[str]:
    warnings: list[str] = []
    leaf_lines = [line for line in fs_lines if not line.summary]
    if len(leaf_lines) < 10:
        warnings.append(
            f"FS extraction has only {len(leaf_lines)} non-summary lines; "
            "detailed mapping may be incomplete"
        )

    totals = {
        line.normalized: line.amount
        for line in fs_lines
        if line.amount is not None
        and line.normalized in {"total general actif", "total general passif"}
    }
    asset = totals.get("total general actif")
    liability = totals.get("total general passif")
    if asset is not None and liability is not None:
        tolerance = max(2.0, 0.001 * max(abs(asset), abs(liability), 1.0))
        if abs(asset - liability) > tolerance:
            warnings.append(
                f"FS is unbalanced: total assets={asset:.2f}, "
                f"total liabilities/equity={liability:.2f}"
            )

    label_amounts: dict[str, set[float]] = defaultdict(set)
    for line in leaf_lines:
        if line.amount is not None and abs(line.amount) > 0.01:
            label_amounts[line.normalized].add(round(line.amount, 2))
    duplicates = [
        label for label, amounts in label_amounts.items() if len(amounts) > 1
    ]
    if duplicates:
        warnings.append(
            "FS contains duplicate detailed labels with different amounts: "
            + ", ".join(sorted(duplicates)[:5])
        )
    return warnings


def reconcile_result(
    rows: list[BGRow], fs_lines: list[FSLine]
) -> dict[str, float] | None:
    pnl_rows = [
        row
        for row in rows
        if account_digits(row.account).startswith(("6", "7"))
    ]
    if not pnl_rows:
        return None
    bg_result = -sum(row.amount for row in pnl_rows)
    candidates = [
        line.amount
        for line in fs_lines
        if line.amount is not None
        and (
            line.normalized == "resultat exercice"
            or line.normalized == "benefice perte"
            or line.normalized.endswith("resultat exercice")
        )
    ]
    if not candidates:
        return None
    fs_result = min(candidates, key=lambda amount: abs(amount - bg_result))
    return {
        "bg_result": round(bg_result, 2),
        "fs_result": round(fs_result, 2),
        "difference": round(fs_result - bg_result, 2),
    }


def write_output(rows: list[BGRow], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "BG_Mapped"
    sheet.append(OUTPUT_HEADERS)
    for row in rows:
        sheet.append([row.account, row.label, row.amount, row.mapping])
    workbook.save(path)


def write_audit_report(stats: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        f"# BG / FS Reconciliation - {stats['ticket']}",
        "",
        f"- Reporting year: {stats.get('reporting_year') or 'not supplied'}",
        f"- Presentation regime: {stats['regime']}",
        (
            f"- Nonzero BG mapped: {stats['matched_nonzero_rows']}/"
            f"{stats['nonzero_rows']}"
        ),
        (
            f"- Nonzero FS lines exactly justified: "
            f"{stats['justified_nonzero_fs_lines']}/"
            f"{stats['nonzero_fs_leaf_lines']}"
        ),
        "",
        "## Quality Warnings",
        "",
    ]
    warnings = stats["quality_warnings"]
    lines.extend(f"- {warning}" for warning in warnings)
    if not warnings:
        lines.append("- None")

    lines.extend(
        [
            "",
            "## FS To BG",
            "",
            "| FS line | FS amount | BG amount | Difference | Status | Categories | Methods | Confidence | Statement conflicts | Accounts |",
            "| --- | ---: | ---: | ---: | --- | --- | --- | ---: | --- | --- |",
        ]
    )
    for item in stats["reconciliation"]:
        lines.append(
            f"| {item['fs_line']} | {item['fs_amount']:.2f} | "
            f"{item['bg_amount']:.2f} | {item['difference']:.2f} | "
            f"{item['status']} | {', '.join(item.get('categories', []))} | "
            f"{', '.join(item.get('mapping_methods', []))} | "
            f"{float(item.get('confidence', 0.0)):.2f} | "
            f"{', '.join(item.get('conflicting_statement_families', []))} | "
            f"{', '.join(item.get('participating_accounts', []))} |"
        )
    result_check = stats.get("result_reconciliation")
    if result_check:
        lines.extend(
            [
                "",
                "## Result Cross-Check",
                "",
                f"- BG-derived result: {result_check['bg_result']:.2f}",
                f"- FS reported result: {result_check['fs_result']:.2f}",
                f"- FS minus BG: {result_check['difference']:.2f}",
            ]
        )

    review_items = [
        item for item in stats["category_differences"] if item["status"] == "review"
    ]
    lines.extend(
        [
            "",
            "## BG To FS Differences",
            "",
            "These are accounting-category expectations, including categories "
            "that strict matching deliberately left blank.",
            "",
            "| Category | FS line | BG expected | FS amount | Difference |",
            "| --- | --- | ---: | ---: | ---: |",
        ]
    )
    for item in review_items:
        lines.append(
            f"| {item['category']} | {item['fs_line']} | "
            f"{item['bg_expected']:.2f} | {item['fs_amount']:.2f} | "
            f"{item['difference']:.2f} |"
        )
    if not review_items:
        lines.append("| None | | | | |")

    if stats["gross_up_adjustments"]:
        lines.extend(["", "## Gross Presentation Inferences", ""])
        for item in stats["gross_up_adjustments"]:
            lines.append(
                f"- {item['fs_lines'][0]} / {item['fs_lines'][1]}: hidden "
                f"counter-balance inferred at "
                f"{item['inferred_hidden_balance']:.2f}. "
                "This is not direct BG justification."
            )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def resolve_input_dir(ticket_or_input: Path) -> Path:
    path = ticket_or_input.resolve()
    if path.is_file():
        path = path.parent
    if (path / "bg_standardized.xlsx").exists():
        return path
    if (path / "Input" / "bg_standardized.xlsx").exists():
        return path / "Input"
    raise FileNotFoundError(f"No bg_standardized.xlsx found under {ticket_or_input}")


def default_output_for(input_dir: Path) -> Path:
    if input_dir.name.lower() == "input":
        return input_dir.parent / "Output" / "BG_Mapped.xlsx"
    return input_dir / "BG_Mapped.xlsx"


def default_fs_output_for(input_dir: Path) -> Path:
    if input_dir.name.lower() == "input":
        return input_dir.parent / "Output" / "FS_Mapped.xlsx"
    return input_dir / "FS_Mapped.xlsx"


def process_ticket(
    ticket_or_input: Path,
    output: Path | None = None,
    fs_override: Path | None = None,
    reporting_year: int | None = None,
    audit_report: Path | None = None,
    fs_output: Path | None = None,
    layout_json: Path | None = None,
    prior_fs_json: Path | None = None,
    bg_override: Path | None = None,
) -> dict[str, Any]:
    input_dir = resolve_input_dir(ticket_or_input)
    bg_path = (
        bg_override.resolve()
        if bg_override
        else input_dir / "bg_standardized.xlsx"
    )
    fs_path = fs_override.resolve() if fs_override else input_dir / "fs_extract_N.json"
    if not fs_path.exists():
        raise FileNotFoundError(f"Missing fs_extract_N.json in {input_dir}")

    rows = load_bg(bg_path)
    fs_lines = load_fs(fs_path)
    if not rows:
        raise ValueError(f"No BG rows found in {bg_path}")
    if not fs_lines:
        raise ValueError(f"No FS lines found in {fs_path}")

    regime, regime_warnings = detect_presentation_regime(
        rows, fs_lines, reporting_year
    )
    mapper = PCGMapper(rows, fs_lines, regime)
    stats = mapper.map()
    output_path = output.resolve() if output else default_output_for(input_dir)
    write_output(rows, output_path)
    layout_path = (
        layout_json.resolve()
        if layout_json
        else input_dir / "fs_tick_layout.json"
    )
    prior_path = (
        prior_fs_json.resolve()
        if prior_fs_json
        else input_dir / "fs_extract_N_1.json"
    )
    fs_output_path = (
        fs_output.resolve() if fs_output else default_fs_output_for(input_dir)
    )
    fs_mapping_rows = build_fs_mapping_rows(
        rows,
        fs_lines,
        mapper,
        stats,
        layout_path if layout_path.exists() else None,
        prior_path if prior_path.exists() else None,
    )
    write_fs_mapping(fs_mapping_rows, fs_output_path)

    nonzero_rows = [row for row in rows if abs(row.amount) > 0.01]
    nonzero_leaf_indexes = {
        index
        for index, line in enumerate(fs_lines)
        if not line.summary and line.amount is not None and abs(line.amount) > 0.01
    }
    matched_nonzero_rows = sum(bool(row.mapping) for row in nonzero_rows)
    justified_nonzero_fs_lines = len(mapper.justified_lines & nonzero_leaf_indexes)
    quality_warnings = fs_quality_warnings(fs_lines) + regime_warnings
    result_check = reconcile_result(rows, fs_lines)
    if result_check is not None:
        tolerance = max(
            2.0,
            0.001
            * max(
                abs(result_check["bg_result"]),
                abs(result_check["fs_result"]),
                1.0,
            ),
        )
        if abs(result_check["difference"]) > tolerance:
            quality_warnings.append(
                "BG-derived result differs from the FS reported result by "
                f"{result_check['difference']:.2f}; review source versions "
                "and the disputed P&L lines"
            )
    if (
        len(nonzero_leaf_indexes) >= 10
        and nonzero_rows
        and matched_nonzero_rows / len(nonzero_rows) < 0.10
    ):
        quality_warnings.append(
            "BG and detailed FS materially fail reconciliation; "
            "check entity, period, extraction, and source completeness"
        )
    unmatched_nonzero_rows = [row for row in nonzero_rows if not row.mapping]
    unmatched_value = sum(abs(row.amount) for row in unmatched_nonzero_rows)
    total_value = sum(abs(row.amount) for row in nonzero_rows)
    if (
        unmatched_nonzero_rows
        and total_value
        and unmatched_value / total_value >= 0.05
    ):
        quality_warnings.append(
            f"{len(unmatched_nonzero_rows)} material BG rows have no compatible "
            "detailed FS line; the extraction may omit visible statement lines"
        )
    if stats["semantic_matches"]:
        quality_warnings.append(
            f"{stats['semantic_matches']} BG rows were mapped by the permissive "
            "semantic fallback despite amount differences or compatibility "
            "warnings; review method, confidence and conflicts in the audit"
        )
    permissive_sum_matches = (
        stats.get("same_statement_permissive_matches", 0)
        + stats.get("cross_statement_permissive_matches", 0)
    )
    if permissive_sum_matches:
        quality_warnings.append(
            f"{permissive_sum_matches} BG rows were mapped by the legacy "
            "equal-sum fallback; cross-statement and low-confidence matches "
            "must be reviewed using the mapper audit metadata"
        )
    if stats["gross_up_adjustments"]:
        quality_warnings.append(
            "Some FS debit/credit gross balances are only inferred from net "
            "standardized BG accounts and are not directly justified"
        )
    if stats.get("fs_framework") == "ifrs_consolidated":
        quality_warnings.append(
            "Consolidated/IFRS presentation signatures were detected; permissive "
            "fallback mappings remain labelled for review and must not be treated "
            "as accounting-backed evidence"
        )
    alignment = stats.get("source_alignment") or {}
    if alignment.get("status") in {"unproven", "weak"}:
        quality_warnings.append(
            "BG/FS source alignment is unproven or weak; permissive mappings are "
            "still emitted for coverage but their method and confidence require "
            "manual review"
        )
    stats.update(
        {
            "ticket": (
                input_dir.parent.name
                if input_dir.name.lower() == "input"
                else input_dir.name
            ),
            "input_dir": str(input_dir),
            "output": str(output_path),
            "fs_output": str(fs_output_path),
            "matched_nonzero_rows": matched_nonzero_rows,
            "nonzero_rows": len(nonzero_rows),
            "justified_nonzero_fs_lines": justified_nonzero_fs_lines,
            "nonzero_fs_leaf_lines": len(nonzero_leaf_indexes),
            "quality_warnings": quality_warnings,
            "reporting_year": reporting_year,
            "result_reconciliation": result_check,
        }
    )
    if audit_report is not None:
        report_path = audit_report.resolve()
        write_audit_report(stats, report_path)
        stats["audit_report"] = str(report_path)
    return stats


def discover_tickets(root: Path) -> list[Path]:
    return sorted({path.parent for path in root.resolve().rglob("bg_standardized.xlsx")})


def print_stats(stats: dict[str, Any]) -> None:
    print(
        f"{stats['ticket']}: mapped nonzero BG "
        f"{stats['matched_nonzero_rows']}/{stats['nonzero_rows']} "
        f"(all rows {stats['matched']}/{stats['rows']}), justified nonzero FS "
        f"{stats['justified_nonzero_fs_lines']}/{stats['nonzero_fs_leaf_lines']}, "
        f"output={stats['output']}, fs_output={stats['fs_output']}"
    )
    for warning in stats["quality_warnings"]:
        print(f"  quality warning: {warning}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Map bg_standardized.xlsx rows to fs_extract_N.json line labels."
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--ticket", type=Path, help="Ticket folder or its Input folder")
    group.add_argument("--batch", type=Path, help="Root containing multiple tickets")
    parser.add_argument("--output", type=Path, help="Single-ticket output path")
    parser.add_argument(
        "--fs-json",
        type=Path,
        help="Optional FS extraction override for a single ticket",
    )
    parser.add_argument(
        "--year",
        type=int,
        help="Reporting year; used when the FS layout has no decisive markers",
    )
    parser.add_argument(
        "--audit-report",
        type=Path,
        help="Optional Markdown FS-to-BG and BG-to-FS reconciliation report",
    )
    parser.add_argument(
        "--fs-output",
        type=Path,
        help="Optional FS-centric reconciliation workbook path",
    )
    parser.add_argument(
        "--layout-json",
        type=Path,
        help="Optional visually extracted PDF tickmark coordinates",
    )
    parser.add_argument(
        "--prior-fs-json",
        type=Path,
        help="Optional independently extracted prior-year FS amounts",
    )
    args = parser.parse_args()

    if args.ticket:
        try:
            stats = process_ticket(
                args.ticket,
                args.output,
                args.fs_json,
                args.year,
                args.audit_report,
                args.fs_output,
                args.layout_json,
                args.prior_fs_json,
            )
        except Exception as exc:
            print(f"{args.ticket}: failed: {exc}", file=sys.stderr)
            return 1
        print_stats(stats)
        return 0

    successes = 0
    skipped = 0
    for input_dir in discover_tickets(args.batch):
        if not (input_dir / "fs_extract_N.json").exists():
            print(f"{input_dir.parent.name}: skipped (missing fs_extract_N.json)")
            skipped += 1
            continue
        try:
            stats = process_ticket(input_dir, reporting_year=args.year)
            print_stats(stats)
            successes += 1
        except Exception as exc:
            print(f"{input_dir.parent.name}: failed: {exc}", file=sys.stderr)
            skipped += 1
    print(f"Batch complete: {successes} processed, {skipped} skipped/failed")
    return 0 if successes else 1


if __name__ == "__main__":
    raise SystemExit(main())
