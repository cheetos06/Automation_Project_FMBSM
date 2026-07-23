from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
SERVICE = ROOT / "services" / "automation-mail-worker"
FS_REVIEW = SERVICE / "fs_review"
for candidate in (str(SERVICE), str(FS_REVIEW)):
    if candidate not in sys.path:
        sys.path.insert(0, candidate)

from balance_cleaner.render import (  # noqa: E402
    _font,
    _text_width,
    _wrap,
    capture_workbook,
    select_capture_columns,
)
from balance_cleaner.run_pipeline import (  # noqa: E402
    _batch_prompt,
    _extract_records,
    _extract_schema,
    _reconcile_account_numbers,
    _schema_prompt,
    _write_workbook,
    chunk_screenshots,
)
from copilot_runtime.runtime import (  # noqa: E402
    _parse_first_json_value,
    _select_response_text,
)
from fmbsm_email_bot.worker import _job_kind  # noqa: E402


class BalanceRendererTests(unittest.TestCase):
    def test_wrapping_uses_pixel_width_without_dropping_visible_text(self) -> None:
        font = _font(False)
        value = "CRED IMPOST ART.1, COMMI DA 1051 A 1063, L. 178/20"
        wrapped = _wrap(value, 398, font)

        self.assertEqual("".join(wrapped.split()), "".join(value.split()))
        self.assertTrue(all(_text_width(font, line) <= 388 for line in wrapped.splitlines()))
        self.assertIn("L.", wrapped)

    def test_renderer_freezes_nonempty_rows_before_chunking(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            workbook_path = root / "input.xlsx"
            workbook = Workbook()
            cover = workbook.active
            cover.title = "Cover"
            cover["A1"] = "Title"
            balance = workbook.create_sheet("Any Balance")
            balance.append(["Compte", "Libellé", "Débit", "Crédit"])
            for row in (2, 99, 100, 101, 199, 200, 201, 205):
                balance.cell(row, 1, f"{row:06d}")
                balance.cell(row, 2, f"Account {row}")
                balance.cell(row, 3, row * 10)
                balance.cell(row, 4, 0)
            # Formatting alone must not inflate the detected last row.
            balance.cell(500, 12).fill = PatternFill("solid", fgColor="FFFF00")
            workbook.save(workbook_path)

            capture = capture_workbook(workbook_path, root / "rendered")

            self.assertEqual(capture.sheet_name, "Any Balance")
            self.assertEqual(capture.last_populated_row, 205)
            self.assertEqual(
                [(item.row_start, item.row_end) for item in capture.screenshots],
                [(1, 205)],
            )
            self.assertEqual(
                capture.screenshots[0].populated_rows,
                (1, 2, 99, 100, 101, 199, 200, 201, 205),
            )
            self.assertEqual(capture.first_populated_column, 1)
            self.assertEqual(capture.last_populated_column, 4)
            self.assertIsNotNone(capture.header_context)
            self.assertEqual(capture.header_context.populated_rows, (1,))
            with Image.open(capture.screenshots[0].path) as image:
                self.assertGreater(image.width, 300)
                self.assertGreater(image.height, 100)

            limited = capture_workbook(
                workbook_path,
                root / "limited",
                rows_per_image=30,
                last_row_limit=120,
            )
            self.assertEqual(limited.last_populated_row, 101)
            self.assertEqual(
                [(item.row_start, item.row_end) for item in limited.screenshots],
                [(1, 101)],
            )

    def test_renderer_freezes_one_compact_column_mask_for_every_image(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            workbook_path = root / "spacers.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "Title only"
            sheet["A2"] = "Second title"
            sheet["C5"] = "Company"
            sheet["D5"] = "Account"
            sheet["E5"] = "One-off spacer heading"
            sheet["F5"] = "Description"
            sheet["Q5"] = "Final balance"
            for row in range(7, 107):
                sheet.cell(row, 3, "BA02")
                sheet.cell(row, 4, f"{row:010d}")
                sheet.cell(row, 6, f"Account {row}")
                sheet.cell(row, 17, float(row))
            sheet.cell(6, 6, "Styled section label").font = Font(bold=True)
            for column in (3, 4, 6, 17):
                sheet.cell(37, column).font = Font(bold=True)
            workbook.save(workbook_path)

            capture = capture_workbook(
                workbook_path,
                root / "rendered",
                rows_per_image=30,
            )

            self.assertEqual(capture.populated_columns, (3, 4, 6, 17))
            self.assertIsNotNone(capture.header_context)
            self.assertEqual(capture.header_context.populated_rows, (5, 6))
            self.assertTrue(
                all(
                    screenshot.populated_columns == capture.populated_columns
                    for screenshot in capture.screenshots
                )
            )
            self.assertTrue(
                all(
                    len(screenshot.populated_rows) <= 30
                    for screenshot in capture.screenshots
                )
            )
            self.assertEqual(capture.screenshots[0].populated_rows[0], 5)
            captured_rows = {
                row
                for screenshot in capture.screenshots
                for row in screenshot.populated_rows
            }
            self.assertNotIn(6, captured_rows)
            self.assertNotIn(37, captured_rows)
            self.assertEqual(
                sum(item.estimated_records for item in capture.screenshots),
                99,
            )

    def test_renderer_removes_only_amountless_duplicate_account_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            workbook_path = root / "duplicates.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Account", "Description", "Debit", "Credit"])
            sheet.append(["1600 0", "Amount-bearing occurrence", None, 283837.12])
            sheet.append(["1600 0", "Explanatory repeat", None, None])
            sheet.append(["1400 0", "First real occurrence", 1250956.78, None])
            sheet.append(["1400 0", "Second real occurrence", 18573.98, None])
            workbook.save(workbook_path)

            capture = capture_workbook(
                workbook_path,
                root / "rendered",
                rows_per_image=30,
            )
            captured_rows = {
                row
                for screenshot in capture.screenshots
                for row in screenshot.populated_rows
            }

            self.assertIn(2, captured_rows)
            self.assertNotIn(3, captured_rows)
            self.assertIn(4, captured_rows)
            self.assertIn(5, captured_rows)
            self.assertEqual(
                sum(item.estimated_records for item in capture.screenshots),
                3,
            )

    def test_schema_columns_are_selected_once_for_every_screenshot(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            workbook_path = root / "wide.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Hierarchy",
                    "Account",
                    "Description",
                    "Current",
                    "Prior",
                    "Variance",
                ]
            )
            for row in range(2, 64):
                sheet.append(
                    [
                        "Section",
                        f"A{row:04d}",
                        f"Account {row}",
                        float(row),
                        float(row - 1),
                        1.0,
                    ]
                )
            workbook.save(workbook_path)

            capture = capture_workbook(
                workbook_path,
                root / "rendered",
                rows_per_image=30,
            )
            selected = select_capture_columns(
                capture,
                (2, 3, 4),
                2,
                root / "selected",
            )

            self.assertEqual(selected.populated_columns, (2, 3, 4))
            self.assertIsNone(selected.header_context)
            self.assertEqual(len(selected.screenshots), len(capture.screenshots))
            self.assertTrue(
                all(
                    item.populated_columns == (2, 3, 4)
                    for item in selected.screenshots
                )
            )
            with (
                Image.open(capture.screenshots[0].path) as full_image,
                Image.open(selected.screenshots[0].path) as selected_image,
            ):
                self.assertLess(selected_image.width, full_image.width)

    def test_copilot_batch_size_is_hard_capped_at_30(self) -> None:
        screenshots = tuple(SimpleNamespace(path=Path(f"{index}.png")) for index in range(65))
        batches = chunk_screenshots(screenshots, 999)  # type: ignore[arg-type]
        self.assertEqual([len(batch) for batch in batches], [30, 30, 5])
        self.assertLessEqual(max(len(batch) for batch in batches), 30)

    def test_copilot_batches_are_limited_by_estimated_record_volume(self) -> None:
        screenshots = tuple(
            SimpleNamespace(
                path=Path(f"{index}.png"),
                estimated_records=60,
            )
            for index in range(5)
        )
        batches = chunk_screenshots(  # type: ignore[arg-type]
            screenshots,
            10,
            150,
        )
        self.assertEqual([len(batch) for batch in batches], [2, 2, 1])

    def test_production_uses_five_target_screenshots_per_request(self) -> None:
        settings = json.loads(
            (
                SERVICE
                / "balance_cleaner"
                / "config"
                / "settings.json"
            ).read_text(encoding="utf-8")
        )
        self.assertEqual(settings["batch_size"], 5)
        self.assertEqual(settings["rows_per_image"], 30)
        self.assertEqual(settings["max_estimated_records_per_request"], 120)

    def test_batch_prompt_retains_the_extraction_instructions(self) -> None:
        screenshot = SimpleNamespace(
            row_start=1,
            row_end=100,
            estimated_records=17,
            candidate_account_numbers=("A100", "A200"),
        )
        capture = SimpleNamespace(sheet_name="Balance")
        base = "EXTRACT THESE ACCOUNTS"
        schema = {
            "account_fields": {
                "account_number_column_position": 2,
                "account_description_column_position": 3,
            },
            "saldo": {"source_column_positions": [4]},
        }
        batch_prompt = _batch_prompt(  # type: ignore[arg-type]
            base,
            capture,
            (screenshot,),
            1,
            2,
            schema,
            (2, 3, 4),
        )
        self.assertTrue(batch_prompt.startswith(base))
        self.assertIn("Excel rows 1-100", batch_prompt)
        self.assertIn("mechanical candidate-row estimate: 17", batch_prompt)
        self.assertIn("navigation and completeness aid", batch_prompt)
        self.assertIn("never extract a title, total, subtotal", batch_prompt)
        self.assertIn('"account_number_candidates":["A100","A200"]', batch_prompt)
        self.assertIn('"source_column_positions":[4]', batch_prompt)
        self.assertIn("only these original 1-based schema column positions", batch_prompt)
        self.assertIn("candidate lists contain 2 amount-bearing account rows", batch_prompt)
        self.assertIn("Never attach one candidate's description or saldo", batch_prompt)
        self.assertNotIn("HEADER/STRUCTURE CONTEXT ONLY", batch_prompt)
        self.assertNotIn("independent verification read", batch_prompt)

    def test_schema_validation_selects_only_account_description_and_saldo_columns(
        self,
    ) -> None:
        schema = {
            "columns": [
                {"position": 1, "role": "hierarchy_or_category"},
                {"position": 2, "role": "account_number"},
                {"position": 3, "role": "account_description"},
                {"position": 4, "role": "current_or_closing_balance"},
                {"position": 5, "role": "comparison_or_prior_period"},
            ],
            "account_fields": {
                "account_number_column_position": 2,
                "account_description_column_position": 3,
            },
            "saldo": {"source_column_positions": [4]},
        }
        schema["saldo"]["valid_account_balance_patterns"] = [  # type: ignore[index]
            {
                "source_column_positions": [4],
                "visible_account_examples": ["A100"],
            }
        ]

        parsed, selected, account_position = _extract_schema(schema, 5)

        self.assertIs(parsed, schema)
        self.assertEqual(selected, (2, 3, 4))
        self.assertEqual(account_position, 2)

    def test_schema_validation_requires_every_visible_column(self) -> None:
        schema = {
            "columns": [{"position": 1}, {"position": 3}],
            "account_fields": {
                "account_number_column_position": 1,
                "account_description_column_position": 2,
            },
            "saldo": {
                "source_column_positions": [3],
                "valid_account_balance_patterns": [
                    {
                        "source_column_positions": [3],
                        "visible_account_examples": ["100"],
                    }
                ],
            },
        }
        with self.assertRaisesRegex(RuntimeError, "every visible column"):
            _extract_schema(schema, 3)

    def test_schema_prompt_adds_candidate_occupancy_without_amounts(self) -> None:
        screenshot = SimpleNamespace(
            candidate_account_numbers=("43 0", "27 0"),
            candidate_numeric_positions=(
                ("43 0", (4, 5)),
                ("27 0", (3, 5)),
            ),
        )

        prompt = _schema_prompt("ANALYZE THE HEADER", screenshot)  # type: ignore[arg-type]

        self.assertIn('"account_number": "43 0"', prompt)
        self.assertIn('"nonempty_numeric_column_positions": [4, 5]', prompt)
        self.assertIn("positions only, never amounts", prompt)
        self.assertNotIn("1649349.79", prompt)

    def test_schema_validation_rejects_horizontal_shift_on_first_account(self) -> None:
        schema = {
            "columns": [
                {"position": 1},
                {"position": 2},
                {"position": 3},
                {"position": 4},
                {"position": 5},
            ],
            "account_fields": {
                "account_number_column_position": 1,
                "account_description_column_position": 2,
            },
            "saldo": {
                "source_column_positions": [3],
                "valid_account_balance_patterns": [
                    {
                        "source_column_positions": [3],
                        "visible_account_examples": ["43 0"],
                    }
                ],
                "worked_visible_examples": [
                    {"account_number": "43 0", "resulting_saldo": 1649349.79}
                ],
            },
        }

        with self.assertRaisesRegex(RuntimeError, "horizontally misaligned"):
            _extract_schema(schema, 5, (("43 0", (4, 5)),))


class BalanceResponseTests(unittest.TestCase):
    def test_reconciles_only_unique_one_character_account_ocr_errors(self) -> None:
        records = [
            {
                "account_number": "Z13641100",
                "account_description": "Account",
                "saldo": 1.0,
            },
            {
                "account_number": "2000740011",
                "account_description": "Account",
                "saldo": 2.0,
            },
            {
                "account_number": "H35150010",
                "account_description": "Account",
                "saldo": 3.0,
            },
        ]

        reconciled, corrections = _reconcile_account_numbers(
            records,
            ("Z136411000", "200074001", "H35150010"),
        )

        self.assertEqual(
            [item["account_number"] for item in reconciled],
            ["Z136411000", "200074001", "H35150010"],
        )
        self.assertEqual(
            corrections,
            [
                {"from": "Z13641100", "to": "Z136411000"},
                {"from": "2000740011", "to": "200074001"},
            ],
        )

    def test_does_not_reconcile_ambiguous_account_ocr_errors(self) -> None:
        records = [
            {
                "account_number": "1002",
                "account_description": "Account",
                "saldo": 1.0,
            }
        ]

        reconciled, corrections = _reconcile_account_numbers(
            records,
            ("1000", "1001"),
        )

        self.assertEqual(reconciled[0]["account_number"], "1002")
        self.assertEqual(corrections, [])

    def test_runtime_accepts_json_arrays_and_prefers_latest_structured_answer(self) -> None:
        value = _parse_first_json_value(
            'Result follows: [{"account_number":"001","account_description":"Cash","saldo":1.5}] done'
        )
        self.assertIsInstance(value, list)
        selected = _select_response_text(
            ["Working...", "[{\"account_number\":\"001\",\"account_description\":\"Cash\",\"saldo\":1.5}]"]
        )
        self.assertTrue(selected.startswith("["))

    def test_records_are_validated_and_reconstructed_with_text_account_numbers(self) -> None:
        records = _extract_records(
            [
                {
                    "account_number": "000100",
                    "account_description": "Caisse",
                    "saldo": "1.234,50",
                },
                {
                    "account_number": "200",
                    "account_description": "Fournisseurs",
                    "saldo": -25,
                },
            ]
        )
        self.assertEqual(records[0]["saldo"], 1234.5)
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "cleaned.xlsx"
            _write_workbook(output, records)
            workbook = load_workbook(output, data_only=True)
            sheet = workbook["Cleaned balance"]
            self.assertEqual(
                tuple(sheet.cell(1, column).value for column in range(1, 4)),
                ("account_number", "account_description", "saldo"),
            )
            self.assertEqual(sheet["A2"].value, "000100")
            self.assertEqual(sheet["C2"].value, 1234.5)
            self.assertEqual(sheet.max_column, 3)
            workbook.close()

    def test_prompt_is_language_and_layout_agnostic(self) -> None:
        prompt = (
            SERVICE
            / "balance_cleaner"
            / "config"
            / "prompts"
            / "balance_extract_v1.md"
        ).read_text(encoding="utf-8")
        self.assertIn("any title, structure, accounting convention, or language", prompt)
        self.assertIn("Débit", prompt)
        self.assertIn("Solde", prompt)
        self.assertIn("dedicated account-number column", prompt)
        self.assertIn("distinguish visually similar characters such as S and 5", prompt)
        self.assertIn("Numeric zero is a real amount", prompt)
        self.assertIn("Treat the first visible row of every TARGET image", prompt)
        self.assertIn("silent completeness pass over every TARGET image", prompt)
        self.assertNotIn("Bilancio di verifica", prompt)

    def test_schema_prompt_explains_later_requests_and_demands_evidence(self) -> None:
        prompt = (
            SERVICE
            / "balance_cleaner"
            / "config"
            / "prompts"
            / "balance_schema_v1.md"
        ).read_text(encoding="utf-8")
        self.assertIn("injected into later extraction requests", prompt)
        self.assertIn("Enumerate ALL visible columns", prompt)
        self.assertIn("visible_examples", prompt)
        self.assertIn("worked_visible_examples", prompt)
        self.assertIn("justification", prompt)
        self.assertIn("first readable row", prompt)
        self.assertIn("valid_account_balance_patterns", prompt)


class BalanceDispatchTests(unittest.TestCase):
    def test_balance_subject_is_dispatched(self) -> None:
        settings = SimpleNamespace(
            subject_prefix="[optimda-extract-dates]",
            fs_subject_prefix="[fs-review]",
            effectif_subject_prefix="[optimda-effectif]",
            balance_subject_prefix="[balance-cleaner]",
        )
        self.assertEqual(
            _job_kind(settings, "[balance-cleaner] client trial balance"),
            "balance_cleaner",
        )


if __name__ == "__main__":
    unittest.main()
